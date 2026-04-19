"""4 Tests für parser/v2/dedupliziere_hauptuebersicht.py"""
import tempfile
from pathlib import Path

import openpyxl
import pytest

from export.excel_writer import MAIN_FIELDS, MAIN_HEADERS, SHEET_HAUPT
from parser.v2.dedupliziere_hauptuebersicht import dedupliziere_hauptuebersicht

ID_COL = MAIN_FIELDS.index("willhaben_id") + 1  # 1-based


def _make_excel(rows: list[dict]) -> Path:
    """Erstellt eine temporäre Excel-Datei mit Hauptübersicht-Sheet und den gegebenen Zeilen."""
    tmp_dir = tempfile.mkdtemp()
    path = Path(tmp_dir) / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_HAUPT
    # Header
    for col_idx, header in enumerate(MAIN_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, field in enumerate(MAIN_FIELDS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(field))
    wb.save(path)
    return path


def _read_ids(path: Path) -> list[str]:
    """Liest alle willhaben_id-Werte aus dem Hauptübersicht-Sheet (ohne Header)."""
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_HAUPT]
    ids = []
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=ID_COL).value
        if val is not None and str(val).strip():
            ids.append(str(val).strip())
    return ids


def _read_row_field(path: Path, wid: str, field: str):
    """Liest den Wert eines Feldes für eine bestimmte willhaben_id."""
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_HAUPT]
    field_col = MAIN_FIELDS.index(field) + 1
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=ID_COL).value
        if val is not None and str(val).strip() == wid:
            return ws.cell(row=row_num, column=field_col).value
    return None


def test_keine_duplikate_unveraendert():
    """2 Zeilen mit verschiedenen IDs → gibt 0 zurück, beide Zeilen bleiben erhalten."""
    path = _make_excel([
        {"willhaben_id": "A001", "event_name": "Konzert A"},
        {"willhaben_id": "B002", "event_name": "Konzert B"},
    ])
    removed = dedupliziere_hauptuebersicht(path)
    assert removed == 0
    ids = _read_ids(path)
    assert set(ids) == {"A001", "B002"}
    assert len(ids) == 2


def test_duplikat_entfernt():
    """2 Zeilen mit gleicher ID → gibt 1 zurück, nur die LETZTE Zeile bleibt."""
    path = _make_excel([
        {"willhaben_id": "DUP001", "event_name": "Erster Eintrag"},
        {"willhaben_id": "DUP001", "event_name": "Zweiter Eintrag"},
    ])
    removed = dedupliziere_hauptuebersicht(path)
    assert removed == 1
    ids = _read_ids(path)
    assert ids == ["DUP001"]
    # Letzter Eintrag soll erhalten bleiben
    event_name = _read_row_field(path, "DUP001", "event_name")
    assert event_name == "Zweiter Eintrag"


def test_drei_duplikate_zwei_entfernt():
    """3 Zeilen mit gleicher ID → gibt 2 zurück, nur die LETZTE Zeile bleibt."""
    path = _make_excel([
        {"willhaben_id": "TRIPLE", "event_name": "Erster"},
        {"willhaben_id": "TRIPLE", "event_name": "Zweiter"},
        {"willhaben_id": "TRIPLE", "event_name": "Dritter"},
    ])
    removed = dedupliziere_hauptuebersicht(path)
    assert removed == 2
    ids = _read_ids(path)
    assert ids == ["TRIPLE"]
    event_name = _read_row_field(path, "TRIPLE", "event_name")
    assert event_name == "Dritter"


def test_leere_datei_keine_fehler():
    """Leeres Sheet (nur Header) → gibt 0 zurück, kein Fehler."""
    path = _make_excel([])
    removed = dedupliziere_hauptuebersicht(path)
    assert removed == 0
    ids = _read_ids(path)
    assert ids == []
