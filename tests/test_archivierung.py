import datetime
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from export.archivierung import archive_expired, archive_aeltere_als
from export.excel_writer import (
    MAIN_FIELDS, MAIN_HEADERS, SHEET_HAUPT, SHEET_ARCHIV, _write_header,
)


def _make_xlsx(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = SHEET_HAUPT
    _write_header(ws_main, MAIN_HEADERS)
    for row in rows:
        ws_main.append([row.get(f, "") for f in MAIN_FIELDS])

    ws_arch = wb.create_sheet(SHEET_ARCHIV)
    _write_header(ws_arch, MAIN_HEADERS)
    wb.save(path)


def test_archive_expired_moves_past_event():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "A1", "event_datum": "2025-01-01", "event_name": "OldEvent"},
            {"willhaben_id": "B1", "event_datum": "2030-12-31", "event_name": "FutureEvent"},
        ])
        count = archive_expired(path, cutoff_date=datetime.date(2026, 1, 1))
        assert count == 1


def test_archive_expired_sets_archiviert_am():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "A1", "event_datum": "2025-01-01"},
        ])
        archive_expired(path, cutoff_date=datetime.date(2026, 1, 1))

        wb = load_workbook(path)
        ws_arch = wb[SHEET_ARCHIV]
        col = MAIN_FIELDS.index("archiviert_am") + 1
        val = ws_arch.cell(row=2, column=col).value
        assert val == datetime.date.today().isoformat()


def test_archive_expired_no_past_events():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "B1", "event_datum": "2099-01-01"},
        ])
        count = archive_expired(path, cutoff_date=datetime.date(2026, 1, 1))
        assert count == 0


def test_archive_expired_default_cutoff_is_today():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "A1", "event_datum": "2020-01-01"},
        ])
        count = archive_expired(path)
        assert count == 1


def test_default_cutoff_ist_minus_30_tage():
    today = datetime.date.today()
    old_date = (today - datetime.timedelta(days=31)).isoformat()
    recent_date = (today - datetime.timedelta(days=29)).isoformat()
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "A1", "event_datum": old_date},
        ])
        count = archive_expired(path)
        assert count == 1

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "B1", "event_datum": recent_date},
        ])
        count = archive_expired(path)
        assert count == 0


def test_archive_aeltere_als_30_tage():
    today = datetime.date.today()
    old_date = (today - datetime.timedelta(days=31)).isoformat()
    recent_date = (today - datetime.timedelta(days=29)).isoformat()
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "A1", "event_datum": old_date},
            {"willhaben_id": "B1", "event_datum": recent_date},
        ])
        count = archive_aeltere_als(path, tage=30)
        assert count == 1


def test_archive_aeltere_als_7_tage():
    today = datetime.date.today()
    old_date = (today - datetime.timedelta(days=8)).isoformat()
    recent_date = (today - datetime.timedelta(days=6)).isoformat()
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "A1", "event_datum": old_date},
            {"willhaben_id": "B1", "event_datum": recent_date},
        ])
        count = archive_aeltere_als(path, tage=7)
        assert count == 1


def test_archive_aeltere_als_null_tage():
    today = datetime.date.today()
    past_date = (today - datetime.timedelta(days=1)).isoformat()
    future_date = (today + datetime.timedelta(days=1)).isoformat()
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "A1", "event_datum": past_date},
            {"willhaben_id": "B1", "event_datum": future_date},
        ])
        count = archive_aeltere_als(path, tage=0)
        assert count == 1
