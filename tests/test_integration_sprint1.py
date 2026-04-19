"""
Integration-Test Dashboard Sprint 1:
Scraper-Dict → finalisiere_lauf() → echtes Excel (34 Spalten)
→ archivierung → Dashboard-Sheet → aggregator
"""
import datetime
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from export.excel_writer import (
    finalisiere_lauf, upsert_events,
    MAIN_FIELDS, MAIN_HEADERS,
    SHEET_DASHBOARD, SHEET_HAUPT, SHEET_REVIEW, SHEET_WATCHLIST, SHEET_ARCHIV,
    DASHBOARD_HEADERS,
)
from export.archivierung import archive_expired
from app.backend.dashboard_aggregator import load_excel, aggregate, _OUTPUT_COLUMNS


_EVENTS = [
    {
        "willhaben_id":          "INT_001",
        "willhaben_link":        "https://willhaben.at/INT_001",
        "event_name":            "Linkin Park",
        "event_datum":           "2030-06-09",
        "venue":                 "Ernst Happel",
        "stadt":                 "Wien",
        "kategorie":             "Stehplatz",
        "anzahl_karten":         2,
        "angebotspreis_gesamt":  250.0,
        "preis_ist_pro_karte":   False,
        "originalpreis_pro_karte": 89.9,
        "ovp_quelle":            "oeticket",
        "ausverkauft":           "nein",
        "verkäufertyp":          "Händler",
        "verkäufername":         "TicketKing",
        "verkäufer_id":          "1234",
        "mitglied_seit":         "01/2018",
        "confidence":            "hoch",
        "eingestellt_am":        "2026-04-19",
        "modell":                "gemma3:27b",
        "pipeline_version":      "v2.0",
        "parse_dauer_ms":        1500,
    },
    {
        "willhaben_id":          "INT_002",
        "willhaben_link":        "https://willhaben.at/INT_002",
        "event_name":            "linkin park",          # Kleinschreibung → gleiche Gruppe
        "event_datum":           "2030-06-09",
        "venue":                 "Happel Stadion",
        "stadt":                 "Wien",
        "kategorie":             "Stehplatz",
        "anzahl_karten":         1,
        "angebotspreis_gesamt":  130.0,
        "preis_ist_pro_karte":   True,
        "verkäufertyp":          "Privat",
        "verkäufername":         "Anna",
        "verkäufer_id":          "5678",
        "mitglied_seit":         "06/2022",
        "confidence":            "hoch",
        "eingestellt_am":        "2026-04-19",
    },
    {
        "willhaben_id":          "INT_003",
        "willhaben_link":        "https://willhaben.at/INT_003",
        "event_name":            "OldConcert",
        "event_datum":           "2020-01-01",            # Vergangenheit → archivieren
        "venue":                 "Gasometer",
        "stadt":                 "Wien",
        "kategorie":             "Sitzplatz",
        "anzahl_karten":         1,
        "angebotspreis_gesamt":  50.0,
        "preis_ist_pro_karte":   True,
        "verkäufertyp":          "Privat",
        "verkäufername":         "Hans",
        "verkäufer_id":          "9999",
        "mitglied_seit":         "01/2015",
        "confidence":            "mittel",
        "eingestellt_am":        "2020-01-01",
    },
]


def test_sheet_order():
    """Dashboard ist erstes Sheet, Alte Veranstaltungen ist letztes."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sprint1.xlsx"
        finalisiere_lauf(_EVENTS, path)
        wb = load_workbook(path)
        assert wb.sheetnames[0]  == SHEET_DASHBOARD
        assert wb.sheetnames[-1] == SHEET_ARCHIV
        assert SHEET_HAUPT     in wb.sheetnames
        assert SHEET_REVIEW    in wb.sheetnames
        assert SHEET_WATCHLIST in wb.sheetnames


def test_hauptuebersicht_has_34_columns():
    """Hauptübersicht hat nach Sprint-2-Erweiterung 42 Spalten (34 + 8 Historien)."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sprint1.xlsx"
        finalisiere_lauf(_EVENTS, path)
        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        assert ws.max_column == 42
        assert len(MAIN_FIELDS) == 42


def test_sprint1_enrichment_fields_written():
    """vertrieb_klasse, venue_normiert und venue_typ werden automatisch befüllt."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sprint1.xlsx"
        finalisiere_lauf(_EVENTS, path)
        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        hdrs = [ws.cell(row=1, column=c).value for c in range(1, 35)]

        vk_col  = hdrs.index("Vertriebsklasse") + 1
        vnorm   = hdrs.index("Venue (normiert)") + 1
        vtyp    = hdrs.index("Venue-Typ") + 1
        vkap    = hdrs.index("Venue-Kapazität") + 1

        # INT_001: Händler → gewerblich, "Ernst Happel" → Ernst-Happel-Stadion
        row2 = {hdrs[c - 1]: ws.cell(row=2, column=c).value for c in range(1, 35)}
        assert row2["Vertriebsklasse"]    == "gewerblich"
        assert row2["Venue (normiert)"]   == "Ernst-Happel-Stadion"
        assert row2["Venue-Typ"]          == "Stadion"
        assert row2["Venue-Kapazität"]    == 51000

        # INT_002: Privat + Kleinschreibung event_name
        row3 = {hdrs[c - 1]: ws.cell(row=3, column=c).value for c in range(1, 35)}
        assert row3["Vertriebsklasse"]    == "privat"
        assert row3["Venue (normiert)"]   == "Ernst-Happel-Stadion"


def test_archivierung_sets_archiviert_am():
    """OldConcert wird archiviert, archiviert_am wird gesetzt."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sprint1.xlsx"
        result = finalisiere_lauf(_EVENTS, path)

        archived = result["archived"]
        assert archived == 1

        wb = load_workbook(path)
        ws_arch = wb[SHEET_ARCHIV]
        col_idx = MAIN_FIELDS.index("archiviert_am") + 1
        val = ws_arch.cell(row=2, column=col_idx).value
        assert val == datetime.date.today().isoformat()


def test_dashboard_sheet_has_aggregated_data():
    """Dashboard-Sheet enthält aggregierte Zeilen mit korrekten Spalten."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sprint1.xlsx"
        finalisiere_lauf(_EVENTS, path)
        wb = load_workbook(path)
        ws_dash = wb[SHEET_DASHBOARD]

        dash_hdrs = [ws_dash.cell(row=1, column=c).value
                     for c in range(1, ws_dash.max_column + 1)]

        assert "Event"            in dash_hdrs
        assert "Händler Avg €/K"  in dash_hdrs
        assert "Top Verkäufer"    in dash_hdrs
        assert "Confidence"       in dash_hdrs
        assert ws_dash.max_row >= 2  # mindestens 1 Datenzeile


def test_aggregator_normalizes_event_name():
    """'Linkin Park' und 'linkin park' landen in derselben Aggregationsgruppe."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sprint1.xlsx"
        finalisiere_lauf(_EVENTS, path)
        df = load_excel(path)
        agg = aggregate(df)

        lp_rows = agg[agg["Event"].str.lower().str.contains("linkin park", na=False)]
        assert len(lp_rows) == 1
        assert lp_rows.iloc[0]["Gesamt_Anzahl"] == 2


def test_aggregator_output_has_all_output_columns():
    """aggregate() gibt alle _OUTPUT_COLUMNS zurück."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sprint1.xlsx"
        finalisiere_lauf(_EVENTS, path)
        df = load_excel(path)
        agg = aggregate(df)
        for col in _OUTPUT_COLUMNS:
            assert col in agg.columns, f"Fehlende Spalte: {col}"


def test_upsert_idempotent():
    """Gleiche ID zweimal eingefügt → inserted=1, updated=1, keine Duplikate."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "sprint1.xlsx"
        r1 = upsert_events([_EVENTS[0]], path)
        r2 = upsert_events([_EVENTS[0]], path)

        assert r1["inserted"] == 1
        assert r2["updated"]  == 1

        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        assert ws.max_row == 2  # Header + 1 Datenzeile, kein Duplikat


def test_column_count_sanity():
    """Sicherheits-Check: MAIN_FIELDS = 42 (34 Sprint-1 + 8 Sprint-2), DASHBOARD_FIELDS = 26."""
    from export.excel_writer import DASHBOARD_FIELDS
    assert len(MAIN_FIELDS)      == 42
    assert len(MAIN_HEADERS)     == 42
    assert len(DASHBOARD_FIELDS) == 26
