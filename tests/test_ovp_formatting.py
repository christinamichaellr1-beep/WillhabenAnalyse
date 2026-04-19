"""Tests for Phase 5: OVP conditional formatting."""
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from export.excel_writer import (
    finalisiere_lauf,
    SHEET_HAUPT,
    MAIN_HEADERS,
    COLOR_OVP_FEHLT,
    COLOR_OVP_EXTRAHIERT,
    COLOR_OVP_KONFLIKT,
)


def _base_event(wid: str, ovp_manuell=None, ovp_final_quelle="extrahiert",
                originalpreis_pro_karte=89.9) -> dict:
    return {
        "willhaben_id": wid,
        "willhaben_link": f"https://willhaben.at/{wid}",
        "event_name": "Test Event",
        "event_datum": "2030-01-01",
        "venue": "Gasometer",
        "stadt": "Wien",
        "kategorie": "Stehplatz",
        "anzahl_karten": 1,
        "angebotspreis_gesamt": 120.0,
        "preis_ist_pro_karte": True,
        "originalpreis_pro_karte": originalpreis_pro_karte,
        "ovp_final_quelle": ovp_final_quelle,
        "ovp_manuell": ovp_manuell,
        "verkäufertyp": "Privat",
        "verkäufername": "Test",
        "verkäufer_id": "999",
        "mitglied_seit": "01/2020",
        "confidence": "hoch",
        "eingestellt_am": "2026-04-19",
    }


def test_ovp_konflikt_cell_is_orange():
    """OVP-Quelle-Zelle ist orange (#FFDDAA) wenn ovp_final_quelle='konflikt'."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        events = [_base_event("F001", ovp_manuell=75.0, ovp_final_quelle="konflikt",
                               originalpreis_pro_karte=100.0)]
        finalisiere_lauf(events, path)

        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        ovp_q_col = MAIN_HEADERS.index("OVP Quelle") + 1
        cell = ws.cell(row=2, column=ovp_q_col)
        assert cell.fill.fgColor.rgb.endswith(COLOR_OVP_KONFLIKT), \
            f"Expected orange fill, got {cell.fill.fgColor.rgb}"


def test_ovp_fehlt_cell_is_red():
    """OVP-Quelle-Zelle ist hellrot (#FFD6D6) wenn kein OVP vorhanden (leere Quelle)."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        events = [_base_event("F002", ovp_manuell=None, ovp_final_quelle=None,
                               originalpreis_pro_karte=None)]
        finalisiere_lauf(events, path)

        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        ovp_q_col = MAIN_HEADERS.index("OVP Quelle") + 1
        cell = ws.cell(row=2, column=ovp_q_col)
        assert cell.fill.fgColor.rgb.endswith(COLOR_OVP_FEHLT), \
            f"Expected red fill, got {cell.fill.fgColor.rgb}"


def test_ovp_manuell_cell_has_no_color_fill():
    """OVP-Quelle-Zelle hat keine Farbfüllung wenn ovp_final_quelle='manuell'."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        events = [_base_event("F003", ovp_manuell=80.0, ovp_final_quelle="manuell",
                               originalpreis_pro_karte=None)]
        finalisiere_lauf(events, path)

        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        ovp_q_col = MAIN_HEADERS.index("OVP Quelle") + 1
        cell = ws.cell(row=2, column=ovp_q_col)
        # No fill should be applied for manuell/beide_übereinstimmend
        rgb = cell.fill.fgColor.rgb
        assert not rgb.endswith(COLOR_OVP_FEHLT), "manuell should not have red fill"
        assert not rgb.endswith(COLOR_OVP_KONFLIKT), "manuell should not have orange fill"
