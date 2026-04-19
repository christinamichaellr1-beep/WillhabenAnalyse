"""
Sprint-2 Integration-Tests + Dashboard-Spalten-Tests.

Phase 4c: _preis_bewegung Unit-Tests + Dashboard-Spalten-Smoke
Phase 5:  3-Tage-Simulation, Inaktiv-Detection (25 Tage), Dashboard aktuell vs historisch
"""
import tempfile
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.backend.dashboard_aggregator import (
    _OUTPUT_COLUMNS,
    _preis_bewegung,
    aggregate,
    load_excel,
)
from export.excel_writer import (
    MAIN_FIELDS,
    SHEET_HAUPT,
    update_hauptuebersicht_mit_historie,
    finalisiere_lauf,
)

# ---------------------------------------------------------------------------
# Fixture-Daten
# ---------------------------------------------------------------------------

_BASE = {
    "willhaben_id":            "H001",
    "willhaben_link":          "https://willhaben.at/H001",
    "event_name":              "Konzert Sprint2",
    "event_datum":             "2026-12-15",
    "venue":                   "Stadthalle Wien",
    "stadt":                   "Wien",
    "kategorie":               "Stehplatz",
    "anzahl_karten":           2,
    "angebotspreis_gesamt":    100.0,
    "preis_ist_pro_karte":     False,
    "originalpreis_pro_karte": 45.0,
    "ovp_quelle":              "oeticket",
    "ausverkauft":             "nein",
    "watchlist":               "nein",
    "confidence":              "hoch",
    "confidence_grund":        "OVP direkt",
    "verkäufertyp":            "Privat",
    "verkäufername":           "Test User",
    "verkäufer_id":            "V001",
    "mitglied_seit":           "01/2020",
    "modell":                  "gemma3:27b",
    "pipeline_version":        "v2.0",
    "parse_dauer_ms":          800,
    "eingestellt_am":          "2026-03-01",
}

_B = {**_BASE, "willhaben_id": "H002", "verkäufername": "User B", "verkäufer_id": "V002"}


# ---------------------------------------------------------------------------
# Phase 4c: _preis_bewegung Unit-Tests
# ---------------------------------------------------------------------------

def test_preis_bewegung_sinkend():
    """Preis fällt um >=3% → 📉"""
    assert "📉" in _preis_bewegung(45.0, 50.0)


def test_preis_bewegung_steigend():
    """Preis steigt um >=3% → 📈"""
    assert "📈" in _preis_bewegung(55.0, 50.0)


def test_preis_bewegung_stabil():
    """Preis ändert sich <3% → ➡ stabil"""
    assert _preis_bewegung(51.0, 50.0) == "➡ stabil"


def test_preis_bewegung_kein_historisch():
    """Kein historischer Preis → ➡ stabil"""
    assert _preis_bewegung(50.0, None) == "➡ stabil"


def test_preis_bewegung_prozent_korrekt():
    """Prozentsatz wird im String angezeigt: -10% bei 45 vs 50"""
    result = _preis_bewegung(45.0, 50.0)
    assert "-10" in result


def test_preis_bewegung_plus_vorzeichen_bei_steigerung():
    """Steigerung zeigt + im Prozentsatz: +20% bei 60 vs 50"""
    result = _preis_bewegung(60.0, 50.0)
    assert "+20" in result


def test_dashboard_output_hat_alle_sprint2_spalten():
    """aggregate() DataFrame enthält alle 6 Sprint-2-Spalten."""
    df = pd.DataFrame([{
        "event_name":         "Test",
        "event_datum":        "2026-12-01",
        "kategorie":          "Stehplatz",
        "anbieter_typ":       "Privat",
        "preis_pro_karte":    50.0,
        "originalpreis_pro_karte": None,
        "confidence":         "hoch",
        "verkäufername":      "Anna",
        "vertrieb_klasse":    "privat",
        "venue_normiert":     None,
        "venue_kapazität":    None,
        "venue_typ":          None,
        "zuletzt_gesehen":    date.today().isoformat(),
        "status":             "aktiv",
        "preis_aktuell":      50.0,
        "preis_vor_7_tagen":  None,
        "venue":              None,
        "stadt":              None,
    }])
    agg = aggregate(df)
    for col in ["Aktiv_7_Tage", "Privat_Avg_Aktuell", "Privat_Avg_Historisch",
                "Haendler_Avg_Aktuell", "Haendler_Avg_Historisch", "Preis_Bewegung"]:
        assert col in agg.columns, f"Fehlt: {col}"


# ---------------------------------------------------------------------------
# Phase 5: 3-Tage-Simulation
# ---------------------------------------------------------------------------

def test_drei_laeufe_scan_anzahl():
    """3 Läufe mit gleicher ID → scan_anzahl akkumuliert auf 3."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        d1, d2, d3 = date(2026, 4, 1), date(2026, 4, 8), date(2026, 4, 15)

        update_hauptuebersicht_mit_historie([_BASE], path, scan_datum=d1)
        update_hauptuebersicht_mit_historie([_BASE], path, scan_datum=d2)
        update_hauptuebersicht_mit_historie([_BASE], path, scan_datum=d3)

        ws = load_workbook(path)[SHEET_HAUPT]
        col = MAIN_FIELDS.index("scan_anzahl") + 1
        assert ws.cell(row=2, column=col).value == 3


def test_drei_laeufe_erstmals_gesehen_unveraendert():
    """erstmals_gesehen bleibt beim ersten Lauf-Datum, wird nie überschrieben."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        update_hauptuebersicht_mit_historie([_BASE], path, scan_datum=date(2026, 4, 1))
        update_hauptuebersicht_mit_historie([_BASE], path, scan_datum=date(2026, 4, 8))

        ws = load_workbook(path)[SHEET_HAUPT]
        col = MAIN_FIELDS.index("erstmals_gesehen") + 1
        assert ws.cell(row=2, column=col).value == "2026-04-01"


def test_drei_laeufe_preis_progression_nach_7_tagen():
    """Lauf 1: Preis 50. Lauf 2 (+7 Tage): Preis 45 → preis_vor_7_tagen=50, preis_aktuell=45."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        e1 = {**_BASE, "angebotspreis_gesamt": 100.0}  # pro_karte = 50
        e2 = {**_BASE, "angebotspreis_gesamt":  90.0}  # pro_karte = 45

        update_hauptuebersicht_mit_historie([e1], path, scan_datum=date(2026, 4, 1))
        update_hauptuebersicht_mit_historie([e2], path, scan_datum=date(2026, 4, 8))

        ws = load_workbook(path)[SHEET_HAUPT]
        akt_col  = MAIN_FIELDS.index("preis_aktuell") + 1
        vor7_col = MAIN_FIELDS.index("preis_vor_7_tagen") + 1
        assert ws.cell(row=2, column=akt_col).value == 45.0
        assert ws.cell(row=2, column=vor7_col).value == 50.0


def test_drei_laeufe_preis_kein_historisch_unter_7_tagen():
    """Preisänderung innerhalb von 7 Tagen → preis_vor_7_tagen bleibt leer."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        e1 = {**_BASE, "angebotspreis_gesamt": 100.0}
        e2 = {**_BASE, "angebotspreis_gesamt":  90.0}

        update_hauptuebersicht_mit_historie([e1], path, scan_datum=date(2026, 4, 1))
        update_hauptuebersicht_mit_historie([e2], path, scan_datum=date(2026, 4, 3))  # nur 2 Tage

        ws = load_workbook(path)[SHEET_HAUPT]
        vor7_col = MAIN_FIELDS.index("preis_vor_7_tagen") + 1
        val = ws.cell(row=2, column=vor7_col).value
        assert val in (None, "")


# ---------------------------------------------------------------------------
# Phase 5: Inaktiv-Detection
# ---------------------------------------------------------------------------

def test_inaktiv_nach_25_tagen():
    """H002 nach 25 Tagen nicht mehr im Scan → status=inaktiv."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        update_hauptuebersicht_mit_historie([_BASE, _B], path, scan_datum=date(2026, 4, 1))
        # Nur H001 im zweiten Lauf — H002 fehlt 25 Tage → inaktiv
        update_hauptuebersicht_mit_historie([_BASE], path, scan_datum=date(2026, 4, 26))

        ws = load_workbook(path)[SHEET_HAUPT]
        id_col  = MAIN_FIELDS.index("willhaben_id") + 1
        sts_col = MAIN_FIELDS.index("status") + 1
        status_map = {
            str(ws.cell(row=r, column=id_col).value): ws.cell(row=r, column=sts_col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=id_col).value not in (None, "")
        }
        assert status_map["H001"] == "aktiv"
        assert status_map["H002"] == "inaktiv"


def test_aktiv_innerhalb_grace_period():
    """H002 nach nur 10 Tagen nicht gesehen → noch aktiv (Grace = 21 Tage)."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        update_hauptuebersicht_mit_historie([_BASE, _B], path, scan_datum=date(2026, 4, 1))
        update_hauptuebersicht_mit_historie([_BASE], path, scan_datum=date(2026, 4, 11))

        ws = load_workbook(path)[SHEET_HAUPT]
        id_col  = MAIN_FIELDS.index("willhaben_id") + 1
        sts_col = MAIN_FIELDS.index("status") + 1
        status_map = {
            str(ws.cell(row=r, column=id_col).value): ws.cell(row=r, column=sts_col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=id_col).value not in (None, "")
        }
        assert status_map["H002"] == "aktiv"


# ---------------------------------------------------------------------------
# Phase 5: Dashboard aktuell vs historisch
# ---------------------------------------------------------------------------

def test_dashboard_preis_bewegung_bei_preissenkung():
    """Nach Preissenkung >3% zeigt Preis_Bewegung 📉 im Dashboard."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        e1 = {**_BASE, "angebotspreis_gesamt": 100.0}  # 50 €/K
        e2 = {**_BASE, "angebotspreis_gesamt":  90.0}  # 45 €/K

        update_hauptuebersicht_mit_historie([e1], path, scan_datum=date(2026, 4, 1))
        update_hauptuebersicht_mit_historie([e2], path, scan_datum=date(2026, 4, 8))

        df = load_excel(path)
        agg = aggregate(df)
        row = agg[agg["Event"].str.contains("Sprint2", na=False)].iloc[0]
        assert "📉" in str(row["Preis_Bewegung"])


def test_dashboard_aktiv_7_tage_zaehlt_korrekt():
    """Aktiv_7_Tage: nur Listings mit status=aktiv + zuletzt_gesehen innerhalb 7 Tage."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        # Lauf 1: beide anlegen
        update_hauptuebersicht_mit_historie([_BASE, _B], path, scan_datum=date(2026, 4, 1))
        # Lauf 2 (25 Tage): nur H001 — H002 wird inaktiv
        update_hauptuebersicht_mit_historie([_BASE], path, scan_datum=date(2026, 4, 26))

        df = load_excel(path)
        agg = aggregate(df)
        row = agg[agg["Event"].str.contains("Sprint2", na=False)].iloc[0]
        # H001: aktiv + zuletzt_gesehen=2026-04-26 (heute-7 ≤ 2026-04-26 → aktuell)
        # H002: inaktiv → nicht gezählt
        assert row["Aktiv_7_Tage"] == 1


def test_dashboard_privat_avg_aktuell_vs_historisch():
    """Privat_Avg_Aktuell nutzt nur Listings mit zuletzt_gesehen innerhalb 7 Tage."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "s2.xlsx"
        # H001: teuer, zuletzt_gesehen = heute (aktuell)
        # H002: günstig, zuletzt_gesehen = vor 30 Tagen (historisch)
        update_hauptuebersicht_mit_historie([_BASE, _B], path, scan_datum=date(2026, 4, 1))

        # Direkt H002 auf altes Datum setzen (simuliert vergangenen Scan)
        # H001 in neuem Scan mit höherem Preis
        e1_neu = {**_BASE, "angebotspreis_gesamt": 120.0}  # 60 €/K
        update_hauptuebersicht_mit_historie([e1_neu], path, scan_datum=date(2026, 5, 1))
        # H002 wurde seit April 1 nicht mehr gesehen (30 Tage → inaktiv + historisch)

        df = load_excel(path)
        agg = aggregate(df)
        row = agg[agg["Event"].str.contains("Sprint2", na=False)].iloc[0]

        # Privat_Avg_Aktuell: H001 aktuell (60€/K)
        # Privat_Avg_Historisch: H002 historisch (50€/K)
        assert row["Privat_Avg_Aktuell"] == 60.0
        assert row["Privat_Avg_Historisch"] == 50.0
