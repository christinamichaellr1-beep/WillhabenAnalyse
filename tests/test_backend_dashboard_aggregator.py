"""Tests for app/backend/dashboard_aggregator.py"""
import math
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest

from app.backend.dashboard_aggregator import aggregate, export_csv, load_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_df(**overrides) -> pd.DataFrame:
    """Create a minimal sample DataFrame for the 'Angebote' sheet."""
    base = {
        "event_name": ["Linkin Park", "Linkin Park", "Linkin Park"],
        "event_datum": ["2026-06-09", "2026-06-09", "2026-06-09"],
        "kategorie": ["Front-of-Stage", "Front-of-Stage", "Front-of-Stage"],
        "venue": ["Ernst Happel Stadion"] * 3,
        "stadt": ["Wien"] * 3,
        "anbieter_typ": ["Händler", "Händler", "Privat"],
        "preis_pro_karte": [250.0, 300.0, 150.0],
        "originalpreis_pro_karte": [89.9, 89.9, 89.9],
    }
    base.update(overrides)
    return pd.DataFrame(base)


# ---------------------------------------------------------------------------
# Tests: load_excel
# ---------------------------------------------------------------------------

def test_load_excel_returns_empty_on_missing_file(tmp_path):
    result = load_excel(tmp_path / "nonexistent.xlsx")
    assert isinstance(result, pd.DataFrame)
    assert result.empty


def test_load_excel_returns_empty_on_wrong_sheet(tmp_path):
    # Write a real xlsx with a different sheet name to trigger exception
    import openpyxl
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "WrongSheet"
    wb.save(path)
    result = load_excel(path)
    assert isinstance(result, pd.DataFrame)
    assert result.empty


def test_load_excel_reads_hauptuebersicht_sheet(tmp_path):
    """C08 regression: load_excel must read sheet 'Hauptübersicht', not 'Angebote'."""
    import openpyxl
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hauptübersicht"
    ws.append(["Event-Name", "Event-Datum", "Verkäufertyp"])
    ws.append(["Linkin Park", "2026-06-09", "Privat"])
    wb.save(path)
    result = load_excel(path)
    assert not result.empty
    assert len(result) == 1


def test_load_excel_renames_columns(tmp_path):
    """C09 regression: load_excel must rename German display headers to snake_case."""
    import openpyxl
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hauptübersicht"
    ws.append(["Event-Name", "Event-Datum", "Verkäufertyp",
               "Anzahl Karten", "Angebotspreis gesamt",
               "Angebotspreis pro Karte", "Originalpreis pro Karte",
               "Venue", "Stadt", "Kategorie"])
    ws.append(["Linkin Park", "2026-06-09", "Händler",
               2, 200.0, 100.0, 89.9,
               "Ernst Happel", "Wien", "Stehplatz"])
    wb.save(path)
    result = load_excel(path)
    assert "event_name" in result.columns
    assert "anbieter_typ" in result.columns
    assert "preis_pro_karte" in result.columns
    assert "originalpreis_pro_karte" in result.columns
    assert "Verkäufertyp" not in result.columns
    assert "Event-Name" not in result.columns


# ---------------------------------------------------------------------------
# Tests: aggregate
# ---------------------------------------------------------------------------

def test_aggregate_basic_grouping():
    df = _make_df()
    result = aggregate(df)
    assert len(result) == 1
    row = result.iloc[0]
    assert row["Event"] == "Linkin Park"
    assert row["Kategorie"] == "Front-of-Stage"


def test_aggregate_separates_privat_vs_haendler():
    df = _make_df()
    result = aggregate(df)
    row = result.iloc[0]
    # 2 Haendler rows
    assert row["Haendler_Anzahl"] == 2
    assert row["Haendler_Min"] == pytest.approx(250.0)
    assert row["Haendler_Max"] == pytest.approx(300.0)
    assert row["Haendler_Avg"] == pytest.approx(275.0)
    # 1 Privat row
    assert row["Privat_Anzahl"] == 1
    assert row["Privat_Min"] == pytest.approx(150.0)
    assert row["Privat_Avg"] == pytest.approx(150.0)
    assert row["Privat_Max"] == pytest.approx(150.0)


def test_aggregate_computes_marge_pct():
    df = _make_df()
    result = aggregate(df)
    row = result.iloc[0]
    ovp = 89.9
    # Marge_Haendler_Pct = (OVP - 275) / OVP * 100 (negative: Haendler above OVP)
    expected_haendler = (ovp - 275.0) / ovp * 100
    assert row["Marge_Haendler_Pct"] == pytest.approx(expected_haendler, rel=1e-4)
    # Marge_Privat_Pct = (OVP - 150) / OVP * 100
    expected_privat = (ovp - 150.0) / ovp * 100
    assert row["Marge_Privat_Pct"] == pytest.approx(expected_privat, rel=1e-4)


def test_aggregate_handles_missing_ovp():
    """None/missing OVP → NaN marge."""
    df = _make_df()
    df["originalpreis_pro_karte"] = None
    result = aggregate(df)
    row = result.iloc[0]
    assert math.isnan(row["OVP"])
    assert math.isnan(row["Marge_Haendler_Pct"])
    assert math.isnan(row["Marge_Privat_Pct"])


def test_aggregate_none_anbieter_typ_treated_as_privat():
    """None in anbieter_typ should be treated as Privat."""
    df = pd.DataFrame({
        "event_name": ["Test Event"],
        "event_datum": ["2026-01-01"],
        "kategorie": ["Kat A"],
        "venue": ["Venue"],
        "stadt": ["Stadt"],
        "anbieter_typ": [None],
        "preis_pro_karte": [100.0],
        "originalpreis_pro_karte": [80.0],
    })
    result = aggregate(df)
    assert len(result) == 1
    row = result.iloc[0]
    assert row["Privat_Anzahl"] == 1
    assert row["Haendler_Anzahl"] == 0


def test_aggregate_computes_preis_from_gesamt_and_anzahl():
    """If preis_pro_karte missing, compute from angebotspreis_gesamt / anzahl_karten."""
    df = pd.DataFrame({
        "event_name": ["Test Event"],
        "event_datum": ["2026-01-01"],
        "kategorie": ["Kat A"],
        "venue": ["Venue"],
        "stadt": ["Stadt"],
        "anbieter_typ": ["Privat"],
        "angebotspreis_gesamt": [200.0],
        "anzahl_karten": [2],
        "originalpreis_pro_karte": [90.0],
    })
    result = aggregate(df)
    assert len(result) == 1
    row = result.iloc[0]
    assert row["Privat_Avg"] == pytest.approx(100.0)


def test_aggregate_multiple_groups():
    """Multiple event/kategorie combinations produce multiple rows."""
    df = pd.DataFrame({
        "event_name": ["Event A", "Event A", "Event B"],
        "event_datum": ["2026-01-01", "2026-01-01", "2026-02-01"],
        "kategorie": ["Kat 1", "Kat 2", "Kat 1"],
        "venue": ["V"] * 3,
        "stadt": ["S"] * 3,
        "anbieter_typ": ["Privat", "Privat", "Privat"],
        "preis_pro_karte": [100.0, 200.0, 150.0],
        "originalpreis_pro_karte": [80.0, 80.0, 80.0],
    })
    result = aggregate(df)
    assert len(result) == 3


def test_aggregate_empty_df_returns_empty():
    result = aggregate(pd.DataFrame())
    assert isinstance(result, pd.DataFrame)
    assert result.empty


# ---------------------------------------------------------------------------
# Tests: export_csv
# ---------------------------------------------------------------------------

def test_export_csv_creates_file(tmp_path):
    df = _make_df()
    agg = aggregate(df)
    out_path = tmp_path / "dashboard.csv"
    export_csv(agg, out_path)
    assert out_path.exists()
    content = out_path.read_text(encoding="utf-8")
    assert "Event" in content
    assert "Linkin Park" in content


def test_export_csv_utf8_encoding(tmp_path):
    """CSV must handle German characters (Ö/ä/etc) without errors."""
    df = pd.DataFrame({"Event": ["Österreich Konzert"], "OVP": [99.0]})
    out_path = tmp_path / "out.csv"
    export_csv(df, out_path)
    content = out_path.read_text(encoding="utf-8")
    assert "Österreich" in content


# ---------------------------------------------------------------------------
# Tests: Sprint-1 neue Features
# ---------------------------------------------------------------------------

def test_aggregate_gesamt_anzahl():
    df = _make_df()
    result = aggregate(df)
    assert result.iloc[0]["Gesamt_Anzahl"] == 3


def test_aggregate_marge_eur():
    """Marge_EUR = avg - OVP (absoluter Aufschlag)."""
    df = _make_df()
    result = aggregate(df)
    row = result.iloc[0]
    ovp = 89.9
    # Haendler_Avg = (250+300)/2 = 275
    assert row["Marge_Haendler_EUR"] == pytest.approx(275.0 - ovp, abs=0.01)
    # Privat_Avg = 150
    assert row["Marge_Privat_EUR"] == pytest.approx(150.0 - ovp, abs=0.01)


def test_aggregate_top_verkaeufer():
    df = pd.DataFrame({
        "event_name":              ["Event X"] * 4,
        "event_datum":             ["2026-06-01"] * 4,
        "kategorie":               ["Stehplatz"] * 4,
        "venue":                   ["Wien"] * 4,
        "stadt":                   ["Wien"] * 4,
        "anbieter_typ":            ["Händler"] * 4,
        "preis_pro_karte":         [100.0, 110.0, 120.0, 130.0],
        "originalpreis_pro_karte": [80.0] * 4,
        "verkäufername":           ["TicketKing", "TicketKing", "OtherShop", "TicketKing"],
    })
    result = aggregate(df)
    row = result.iloc[0]
    assert row["Top_Verkaeufer"] == "TicketKing"
    assert row["Top_Verkaeufer_Anzahl"] == 3


def test_aggregate_confidence_modal():
    df = pd.DataFrame({
        "event_name":              ["Event Y"] * 3,
        "event_datum":             ["2026-07-01"] * 3,
        "kategorie":               ["VIP"] * 3,
        "venue":                   ["V"] * 3,
        "stadt":                   ["S"] * 3,
        "anbieter_typ":            ["Privat"] * 3,
        "preis_pro_karte":         [100.0, 110.0, 120.0],
        "originalpreis_pro_karte": [80.0] * 3,
        "confidence":              ["hoch", "hoch", "niedrig"],
    })
    result = aggregate(df)
    assert result.iloc[0]["Confidence_Modal"] == "hoch"


def test_aggregate_venue_normiert_passthrough():
    df = _make_df()
    df["venue_normiert"] = "Ernst-Happel-Stadion"
    df["venue_typ"]      = "Stadion"
    df["venue_kapazität"] = 51000
    result = aggregate(df)
    row = result.iloc[0]
    assert row["Venue_normiert"] == "Ernst-Happel-Stadion"
    assert row["Venue_typ"]      == "Stadion"
    assert row["Venue_kapazität"] == 51000


def test_aggregate_vertrieb_gewerblich_anteil():
    df = pd.DataFrame({
        "event_name":              ["Event Z"] * 4,
        "event_datum":             ["2026-08-01"] * 4,
        "kategorie":               ["Sitzplatz"] * 4,
        "venue":                   ["V"] * 4,
        "stadt":                   ["S"] * 4,
        "anbieter_typ":            ["Händler", "Privat", "Händler", "Privat"],
        "preis_pro_karte":         [100.0, 90.0, 110.0, 85.0],
        "originalpreis_pro_karte": [80.0] * 4,
        "vertrieb_klasse":         ["gewerblich", "privat", "gewerblich", "privat"],
    })
    result = aggregate(df)
    assert result.iloc[0]["Vertrieb_Gewerblich_Anteil_Pct"] == pytest.approx(50.0)


def test_aggregate_normalizes_event_name_for_grouping():
    """Gleicher Event-Name in unterschiedlicher Schreibweise → eine Gruppe."""
    df = pd.DataFrame({
        "event_name":              ["Linkin Park", "LINKIN PARK", "linkin park"],
        "event_datum":             ["2026-06-09"] * 3,
        "kategorie":               ["Stehplatz"] * 3,
        "venue":                   ["Wien"] * 3,
        "stadt":                   ["Wien"] * 3,
        "anbieter_typ":            ["Privat"] * 3,
        "preis_pro_karte":         [100.0, 110.0, 120.0],
        "originalpreis_pro_karte": [80.0] * 3,
    })
    result = aggregate(df)
    assert len(result) == 1
    assert result.iloc[0]["Gesamt_Anzahl"] == 3


def test_aggregate_top_verkaeufer_missing_column():
    """Fehlendes verkäufername-Spalte → None, 0."""
    df = _make_df()
    result = aggregate(df)
    row = result.iloc[0]
    assert row["Top_Verkaeufer"] is None
    assert row["Top_Verkaeufer_Anzahl"] == 0


def test_aggregate_column_map_contains_sprint1_headers():
    from app.backend.dashboard_aggregator import _EXCEL_COLUMN_MAP
    assert "Venue (normiert)"  in _EXCEL_COLUMN_MAP
    assert "Venue-Kapazität"   in _EXCEL_COLUMN_MAP
    assert "Venue-Typ"         in _EXCEL_COLUMN_MAP
    assert "Vertriebsklasse"   in _EXCEL_COLUMN_MAP
    assert "Eingestellt am"    in _EXCEL_COLUMN_MAP
    assert "Confidence"        in _EXCEL_COLUMN_MAP
    assert "Verkäufername"     in _EXCEL_COLUMN_MAP


def test_aggregate_output_columns_complete():
    """aggregate() gibt immer alle _OUTPUT_COLUMNS zurück."""
    from app.backend.dashboard_aggregator import _OUTPUT_COLUMNS
    df = _make_df()
    result = aggregate(df)
    for col in _OUTPUT_COLUMNS:
        assert col in result.columns, f"Fehlende Spalte: {col}"
