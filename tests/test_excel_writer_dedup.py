"""2 Tests für den Deduplizierungs-Bug-Fix in export/excel_writer.py
(update_hauptuebersicht_mit_historie dedupliziert den Input nach willhaben_id)."""
import datetime
import tempfile
from pathlib import Path

import openpyxl
import pytest

from export.excel_writer import MAIN_FIELDS, SHEET_HAUPT, update_hauptuebersicht_mit_historie

ID_COL = MAIN_FIELDS.index("willhaben_id") + 1  # 1-based


def _base_event(wid: str, **kwargs) -> dict:
    ev = {
        "willhaben_id": wid,
        "willhaben_link": f"https://willhaben.at/{wid}",
        "event_name": "Testkonzert",
        "event_datum": "2026-12-31",
        "venue": "Stadthalle Wien",
        "stadt": "Wien",
        "kategorie": "Stehplatz",
        "anzahl_karten": 2,
        "angebotspreis_gesamt": 100.0,
        "preis_ist_pro_karte": False,
        "originalpreis_pro_karte": 50.0,
        "ovp_quelle": "oeticket",
        "ausverkauft": "nein",
        "confidence": "hoch",
        "confidence_grund": None,
        "verkäufertyp": "Privat",
        "verkäufername": "Test User",
        "verkäufer_id": "V001",
        "mitglied_seit": "01/2020",
        "modell": "gemma3:27b",
        "pipeline_version": "v2.0",
        "parse_dauer_ms": 500,
    }
    ev.update(kwargs)
    return ev


def _count_rows_with_id(path: Path, wid: str) -> int:
    """Zählt Zeilen im Hauptübersicht-Sheet mit der gegebenen willhaben_id."""
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_HAUPT]
    count = 0
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=ID_COL).value
        if val is not None and str(val).strip() == wid:
            count += 1
    return count


def _read_field(path: Path, wid: str, field: str):
    """Liest den Wert eines Feldes für eine bestimmte willhaben_id."""
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_HAUPT]
    field_col = MAIN_FIELDS.index(field) + 1
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=ID_COL).value
        if val is not None and str(val).strip() == wid:
            return ws.cell(row=row_num, column=field_col).value
    return None


def test_dedup_in_batch_gleiche_id_doppelt():
    """Events-Liste mit 2 Events gleicher ID → nur 1 Zeile wird geschrieben."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        events = [
            _base_event("SAME001"),
            _base_event("SAME001"),
        ]
        scan_datum = datetime.date(2026, 4, 19)
        stats = update_hauptuebersicht_mit_historie(events, path, scan_datum=scan_datum)

        # Nur 1 inserted (kein Duplikat)
        assert stats["inserted"] == 1
        assert stats["updated"] == 0
        # Nur 1 Zeile mit dieser ID in der Datei
        assert _count_rows_with_id(path, "SAME001") == 1


def test_dedup_letzte_gewinnt():
    """Events [id1_v1, id1_v2] → v2-Daten (letztes Vorkommen) werden in Sheet geschrieben."""
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        events = [
            _base_event("LAST001", event_name="Erster Eintrag", angebotspreis_gesamt=100.0),
            _base_event("LAST001", event_name="Zweiter Eintrag", angebotspreis_gesamt=200.0),
        ]
        scan_datum = datetime.date(2026, 4, 19)
        update_hauptuebersicht_mit_historie(events, path, scan_datum=scan_datum)

        # Nur 1 Zeile mit dieser ID
        assert _count_rows_with_id(path, "LAST001") == 1
        # Letzter Eintrag soll gewonnen haben
        event_name = _read_field(path, "LAST001", "event_name")
        assert event_name == "Zweiter Eintrag"
