"""Tests für die 4 neuen Spalten in excel_writer.py."""
import datetime
import tempfile
from pathlib import Path


NEW_FIELDS = ["confidence_grund", "modell", "pipeline_version", "parse_dauer_ms"]
NEW_HEADERS = ["Confidence-Grund", "Modell", "Pipeline-Version", "Parse-Dauer ms"]


def test_main_columns_contains_new_fields():
    from export.excel_writer import MAIN_COLUMNS
    fields = [f for f, _ in MAIN_COLUMNS]
    for field in NEW_FIELDS:
        assert field in fields, f"Feld '{field}' fehlt in MAIN_COLUMNS"


def test_main_headers_contains_new_headers():
    from export.excel_writer import MAIN_COLUMNS
    headers = [h for _, h in MAIN_COLUMNS]
    for header in NEW_HEADERS:
        assert header in headers, f"Header '{header}' fehlt in MAIN_COLUMNS"


def test_new_columns_at_end_of_main_columns():
    """Neue Spalten müssen NACH den bestehenden 24 Spalten kommen (backward-compat)."""
    from export.excel_writer import MAIN_COLUMNS
    fields = [f for f, _ in MAIN_COLUMNS]
    for field in NEW_FIELDS:
        assert fields.index(field) >= 24, f"Feld '{field}' muss nach Spalte 24 sein"


def test_upsert_writes_new_fields_to_excel():
    from export.excel_writer import upsert_events, MAIN_FIELDS
    import openpyxl

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        event = {
            "willhaben_id": "TEST001",
            "willhaben_link": "https://willhaben.at/test",
            "event_name": "Testkonzert",
            "event_datum": "2026-12-31",
            "confidence": "hoch",
            "confidence_grund": None,
            "modell": "gemma3:27b",
            "pipeline_version": "v2.0",
            "parse_dauer_ms": 1234,
            "verkäufertyp": "Privat",
            "verkäufername": "Test",
            "verkäufer_id": "123",
            "mitglied_seit": "01/2020",
            "kategorie": "Stehplatz",
            "anzahl_karten": 2,
            "angebotspreis_gesamt": 100.0,
            "preis_ist_pro_karte": False,
            "originalpreis_pro_karte": 50.0,
            "preis_roh": "100 €",
        }
        stats = upsert_events([event], path)
        assert stats["inserted"] == 1

        wb = openpyxl.load_workbook(path)
        ws = wb["Hauptübersicht"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        assert "Modell" in headers
        assert "Pipeline-Version" in headers
        assert "Parse-Dauer ms" in headers
        assert "Confidence-Grund" in headers

        # Werte in Zeile 2 prüfen
        modell_col = headers.index("Modell") + 1
        assert ws.cell(2, modell_col).value == "gemma3:27b"

        version_col = headers.index("Pipeline-Version") + 1
        assert ws.cell(2, version_col).value == "v2.0"

        dauer_col = headers.index("Parse-Dauer ms") + 1
        assert ws.cell(2, dauer_col).value == 1234


def test_sprint1_columns_present():
    """Excel Hauptübersicht muss 34 Spalten haben (28 alt + 6 Sprint-1)."""
    from export.excel_writer import MAIN_FIELDS
    sprint1 = ["eingestellt_am", "vertrieb_klasse", "venue_normiert",
               "venue_kapazität", "venue_typ", "archiviert_am"]
    for f in sprint1:
        assert f in MAIN_FIELDS, f"Fehlendes Sprint-1-Feld: {f}"
    assert len(MAIN_FIELDS) == 52  # 34 Sprint-1 + 8 Sprint-2 Historien-Spalten + 5 Verifikations-Spalten + 5 OVP-Pflege


def test_alte_veranstaltungen_sheet_name():
    """Workbook nutzt 'Archiv' als Sheet-Name."""
    import tempfile
    from pathlib import Path
    from export.excel_writer import upsert_events
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        upsert_events([], path)
        wb = load_workbook(path)
        assert "Archiv" in wb.sheetnames
        assert wb.sheetnames[0] == "Dashboard"


def test_finalisiere_lauf_returns_stats():
    """finalisiere_lauf führt upsert + archivierung + Dashboard in einem Aufruf aus."""
    import tempfile
    from pathlib import Path
    from export.excel_writer import finalisiere_lauf
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        events = [{
            "willhaben_id": "FL_001",
            "willhaben_link": "https://willhaben.at/fl001",
            "event_name": "FinTest",
            "event_datum": "2030-06-01",
            "venue": "Gasometer",
            "stadt": "Wien",
            "kategorie": "Stehplatz",
            "anzahl_karten": 2,
            "angebotspreis_gesamt": 100.0,
            "preis_ist_pro_karte": False,
            "verkäufertyp": "Privat",
            "verkäufername": "Hans",
            "verkäufer_id": "123",
            "mitglied_seit": "01/2020",
            "confidence": "hoch",
        }]
        result = finalisiere_lauf(events, path)
        assert result["inserted"] == 1
        assert "archived" in result

        wb = load_workbook(path)
        assert wb.sheetnames[0] == "Dashboard"
        assert "Archiv" in wb.sheetnames


def test_upsert_populates_enrichment_fields():
    """upsert_events befüllt vertrieb_klasse und venue_normiert automatisch."""
    import tempfile
    from pathlib import Path
    from export.excel_writer import upsert_events, MAIN_FIELDS
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        events = [{
            "willhaben_id": "ENRICH_01",
            "willhaben_link": "https://willhaben.at/e01",
            "event_name": "EnrichTest",
            "event_datum": "2030-01-01",
            "venue": "Gasometer",
            "stadt": "Wien",
            "kategorie": "Stehplatz",
            "anzahl_karten": 2,
            "angebotspreis_gesamt": 100.0,
            "preis_ist_pro_karte": False,
            "verkäufertyp": "Händler",
            "verkäufername": "TicketShop",
            "verkäufer_id": "999",
            "mitglied_seit": "01/2020",
            "confidence": "hoch",
        }]
        upsert_events(events, path)
        wb = load_workbook(path)
        ws = wb["Hauptübersicht"]
        hdrs = [ws.cell(row=1, column=c).value for c in range(1, 35)]

        vk_col  = hdrs.index("Vertriebsklasse") + 1
        vnorm   = hdrs.index("Venue (normiert)") + 1
        vtyp    = hdrs.index("Venue-Typ") + 1

        assert ws.cell(row=2, column=vk_col).value == "gewerblich"
        assert ws.cell(row=2, column=vnorm).value  == "Gasometer Wien"
        assert ws.cell(row=2, column=vtyp).value   == "Halle"


def test_old_data_compatible_new_columns_empty():
    """Bestehende Zeilen (v1) bekommen leere neue Spalten — kein Crash."""
    from export.excel_writer import upsert_events
    import openpyxl

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        event_v1 = {
            "willhaben_id": "V1_001",
            "event_name": "Altes Event",
            "event_datum": "2026-11-01",
            "confidence": "hoch",
            "verkäufertyp": "Privat",
            "verkäufername": "Old",
            "verkäufer_id": "1",
            "mitglied_seit": "01/2018",
            "kategorie": "Unbekannt",
            "anzahl_karten": 1,
            "angebotspreis_gesamt": 50.0,
            "preis_ist_pro_karte": True,
            "originalpreis_pro_karte": None,
            "preis_roh": "50 €",
            # v2-Felder fehlen absichtlich
        }
        stats = upsert_events([event_v1], path)
        assert stats["inserted"] == 1
        # Kein Crash = Test bestanden
