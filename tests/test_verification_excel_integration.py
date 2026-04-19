"""Tests for verification/excel_integration.py (Phase 5)."""
from __future__ import annotations
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from export.excel_writer import MAIN_FIELDS, MAIN_HEADERS, SHEET_HAUPT, _write_header
from verification.clients.base import EventCandidate
from verification.excel_integration import (
    _result_to_verif_fields,
    rebuild_nicht_verifiziert_sheet,
    write_verif_result,
)
from verification.matcher import MatchResult
from verification.orchestrator import VerifStatus, VerificationResult

# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _make_xlsx(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_HAUPT
    _write_header(ws, MAIN_HEADERS)
    for row in rows:
        ws.append([row.get(f, "") for f in MAIN_FIELDS])
    wb.save(path)


def _make_result(status: VerifStatus, best_match: MatchResult | None = None, sources_confirmed=None) -> VerificationResult:
    return VerificationResult(
        status=status,
        best_match=best_match,
        sources_checked=[],
        sources_confirmed=sources_confirmed or [],
        verif_datum="2026-04-19",
    )


def _make_match(event_name: str, score: float) -> MatchResult:
    candidate = EventCandidate(event_name=event_name, source="musicbrainz")
    return MatchResult(
        candidate=candidate,
        name_score=score,
        date_score=0.0,
        city_score=0.0,
        total_score=score,
    )


# ---------------------------------------------------------------------------
# Test 1: write_verif_result updates an existing row
# ---------------------------------------------------------------------------

def test_write_verif_result_updates_row():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [{"willhaben_id": "123", "event_name": "Linkin Park"}])

        result = _make_result(
            VerifStatus.VERIFIED,
            best_match=_make_match("Linkin Park Wien", 0.85),
            sources_confirmed=["musicbrainz", "wikidata"],
        )

        found = write_verif_result(path, "123", result)
        assert found is True

        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        verif_status_col = MAIN_FIELDS.index("verif_status") + 1
        cell_val = ws.cell(row=2, column=verif_status_col).value
        assert cell_val == VerifStatus.VERIFIED.value


# ---------------------------------------------------------------------------
# Test 2: write_verif_result returns False for missing willhaben_id
# ---------------------------------------------------------------------------

def test_write_verif_result_missing_id_returns_false():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [{"willhaben_id": "999", "event_name": "Some Event"}])

        result = _make_result(VerifStatus.UNVERIFIED)
        found = write_verif_result(path, "does-not-exist", result)
        assert found is False


# ---------------------------------------------------------------------------
# Test 3: rebuild_nicht_verifiziert_sheet counts correctly
# ---------------------------------------------------------------------------

def test_rebuild_nicht_verifiziert_counts_correctly():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "1", "event_name": "Verified Event", "verif_status": VerifStatus.VERIFIED.value},
            {"willhaben_id": "2", "event_name": "Unverified Event", "verif_status": VerifStatus.UNVERIFIED.value},
        ])

        count = rebuild_nicht_verifiziert_sheet(path)
        assert count == 1

        wb = load_workbook(path)
        from export.excel_writer import SHEET_NICHT_VERIFIZIERT
        assert SHEET_NICHT_VERIFIZIERT in wb.sheetnames
        ws_nv = wb[SHEET_NICHT_VERIFIZIERT]
        # header row + 1 data row
        assert ws_nv.max_row == 2


# ---------------------------------------------------------------------------
# Test 4: rebuild_nicht_verifiziert_sheet includes rows with empty verif_status
# ---------------------------------------------------------------------------

def test_rebuild_nicht_verifiziert_includes_empty_status():
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_xlsx(path, [
            {"willhaben_id": "10", "event_name": "No Status Event"},   # verif_status = ""
            {"willhaben_id": "11", "event_name": "Likely Event", "verif_status": VerifStatus.LIKELY.value},
        ])

        count = rebuild_nicht_verifiziert_sheet(path)
        # "No Status Event" has empty verif_status → included
        # "Likely Event" has verif_status=wahrscheinlich → NOT included
        assert count == 1

        wb = load_workbook(path)
        from export.excel_writer import SHEET_NICHT_VERIFIZIERT
        ws_nv = wb[SHEET_NICHT_VERIFIZIERT]
        wid_col = MAIN_FIELDS.index("willhaben_id") + 1
        row2_id = ws_nv.cell(row=2, column=wid_col).value
        assert str(row2_id) == "10"


# ---------------------------------------------------------------------------
# Test 5: _result_to_verif_fields extracts verif_name and verif_score correctly
# ---------------------------------------------------------------------------

def test_result_to_verif_fields_with_best_match():
    best = _make_match("Coldplay Wien 2026", 0.92)
    result = VerificationResult(
        status=VerifStatus.LIKELY,
        best_match=best,
        sources_checked=["musicbrainz"],
        sources_confirmed=["musicbrainz"],
        verif_datum="2026-04-19",
    )

    fields = _result_to_verif_fields(result)

    assert fields["verif_status"] == VerifStatus.LIKELY.value
    assert fields["verif_name"] == "Coldplay Wien 2026"
    assert fields["verif_score"] == round(0.92, 3)
    assert fields["verif_quellen"] == "musicbrainz"
    assert fields["verif_datum"] == "2026-04-19"
