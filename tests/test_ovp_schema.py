"""Tests for Phase 1: OVP manual columns schema + migration."""
import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


def _make_test_excel(path: Path, with_ovp_cols: bool = False):
    """Creates a minimal Excel with Hauptübersicht sheet."""
    from export.excel_writer import MAIN_HEADERS, SHEET_HAUPT
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_HAUPT
    headers = MAIN_HEADERS[:-5] if not with_ovp_cols else MAIN_HEADERS  # exclude/include OVP cols
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    wb.save(path)


def test_main_columns_includes_ovp_manual():
    """MAIN_COLUMNS enthält alle 5 neuen OVP-Spalten."""
    from export.excel_writer import MAIN_HEADERS
    assert "OVP manuell €/K" in MAIN_HEADERS
    assert "OVP Anbieter-Link" in MAIN_HEADERS
    assert "OVP Quelle" in MAIN_HEADERS
    assert "OVP manuell eingetragen am" in MAIN_HEADERS
    assert "OVP Notiz" in MAIN_HEADERS


def test_migriere_ovp_spalten_adds_columns():
    """migriere_ovp_spalten() fügt fehlende OVP-Spalten hinzu."""
    from export.excel_writer import migriere_ovp_spalten, SHEET_HAUPT
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_test_excel(path, with_ovp_cols=False)
        added = migriere_ovp_spalten(path)
        assert added >= 5
        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "OVP manuell €/K" in headers
        assert "OVP Anbieter-Link" in headers


def test_migriere_ovp_spalten_idempotent():
    """Zweimaliger Aufruf fügt keine Duplikate hinzu."""
    from export.excel_writer import migriere_ovp_spalten, SHEET_HAUPT
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_test_excel(path, with_ovp_cols=False)
        migriere_ovp_spalten(path)
        added_second = migriere_ovp_spalten(path)
        assert added_second == 0
        wb = load_workbook(path)
        ws = wb[SHEET_HAUPT]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers.count("OVP manuell €/K") == 1


def test_migriere_ovp_spalten_creates_backup():
    """Backup wird vor der Migration erstellt."""
    from export.excel_writer import migriere_ovp_spalten
    import datetime
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "test.xlsx"
        _make_test_excel(path, with_ovp_cols=False)
        migriere_ovp_spalten(path)
        today = datetime.date.today().isoformat()
        backup = Path(tmp) / f"test_pre_ovp_{today}.xlsx"
        assert backup.exists()
