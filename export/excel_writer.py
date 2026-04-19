"""
excel_writer.py

Schreibt/aktualisiert die Willhaben-Analyse-Tabelle (openpyxl).
5 Sheets (in Reihenfolge): Dashboard, Hauptübersicht, Review Queue,
Watchlist-Config, Alte Veranstaltungen.

Update-Logik: Willhaben-ID als Schlüssel → bestehende Zeile updaten, nie duplizieren.
OVP-Felder: einmal gefunden → bleiben bei Updates erhalten (preserve_ovp=True).
finalisiere_lauf() orchestriert upsert → archivierung → Dashboard-Refresh.
"""
import datetime
import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Sheet-Namen
# ---------------------------------------------------------------------------

SHEET_DASHBOARD = "Dashboard"
SHEET_HAUPT     = "Hauptübersicht"
SHEET_REVIEW    = "Review Queue"
SHEET_WATCHLIST = "Watchlist-Config"
SHEET_ARCHIV    = "Alte Veranstaltungen"

# ---------------------------------------------------------------------------
# Spalten-Definitionen (34 Spalten)
# ---------------------------------------------------------------------------

# Interne Feldnamen → Header-Anzeigenamen
MAIN_COLUMNS: list[tuple[str, str]] = [
    ("scan_datum",              "Scan-Datum"),
    ("willhaben_link",          "Willhaben-Link"),
    ("willhaben_id",            "Anzeigen-ID"),
    ("verkäufer_id",            "Verkäufer-ID"),
    ("verkäufername",           "Verkäufername"),
    ("verkäufertyp",            "Verkäufertyp"),
    ("mitglied_seit",           "Mitglied seit"),
    ("event_name",              "Event-Name"),
    ("event_datum",             "Event-Datum"),
    ("venue",                   "Venue"),
    ("stadt",                   "Stadt"),
    ("kategorie",               "Kategorie"),
    ("anzahl_karten",           "Anzahl Karten"),
    ("angebotspreis_gesamt",    "Angebotspreis gesamt"),
    ("preis_ist_pro_karte",     "Preis ist pro Karte"),
    ("angebotspreis_pro_karte", "Angebotspreis pro Karte"),  # berechnet
    ("originalpreis_pro_karte", "Originalpreis pro Karte"),  # farbig
    ("ovp_quelle",              "OVP-Quelle"),
    ("marge_eur",               "Marge €"),              # berechnet
    ("marge_pct",               "Marge %"),              # berechnet
    ("ausverkauft",             "Ausverkauft beim Anbieter"),
    ("watchlist",               "Watchlist"),
    ("confidence",              "Confidence"),
    ("review_nötig",            "Review nötig"),
    # v2.0 — neue Felder (Spalten 25–28); ältere Zeilen lassen diese leer
    ("confidence_grund",        "Confidence-Grund"),
    ("modell",                  "Modell"),
    ("pipeline_version",        "Pipeline-Version"),
    ("parse_dauer_ms",          "Parse-Dauer ms"),
    # Sprint 1 — Spalten 29–34
    ("eingestellt_am",          "Eingestellt am"),
    ("vertrieb_klasse",         "Vertriebsklasse"),
    ("venue_normiert",          "Venue (normiert)"),
    ("venue_kapazität",         "Venue-Kapazität"),
    ("venue_typ",               "Venue-Typ"),
    ("archiviert_am",           "Archiviert am"),
    # Sprint 2 — Historien-Architektur (35–42)
    ("erstmals_gesehen",          "Erstmals gesehen"),
    ("zuletzt_gesehen",           "Zuletzt gesehen"),
    ("status",                    "Status"),
    ("scan_anzahl",               "Scan-Anzahl"),
    ("preis_aktuell",             "Preis aktuell €/K"),
    ("preis_vor_7_tagen",         "Preis vor 7+ Tagen €/K"),
    ("preis_aenderungen_count",   "Preis-Änderungen Count"),
    ("letzte_preisaenderung_am",  "Letzte Preisänderung am"),
]

# Feldnamen als einfache Liste (für Indexzugriff)
MAIN_FIELDS = [f for f, _ in MAIN_COLUMNS]
MAIN_HEADERS = [h for _, h in MAIN_COLUMNS]

REVIEW_COLUMNS: list[tuple[str, str]] = [
    ("willhaben_id",         "Anzeigen-ID"),
    ("willhaben_link",       "Willhaben-Link"),
    ("event_name",           "Event-Name"),
    ("event_datum",          "Event-Datum"),
    ("confidence",           "Confidence"),
    ("confidence_grund",     "Confidence-Grund"),
    ("angebotspreis_gesamt", "Angebotspreis gesamt"),
    ("verkäufertyp",         "Verkäufertyp"),
    ("notiz",                "Notiz"),
]
REVIEW_FIELDS = [f for f, _ in REVIEW_COLUMNS]
REVIEW_HEADERS = [h for _, h in REVIEW_COLUMNS]

ARCHIV_FIELDS = MAIN_FIELDS
ARCHIV_HEADERS = MAIN_HEADERS

WATCHLIST_COLUMNS: list[tuple[str, str]] = [
    ("event_name",  "Event-Name"),
    ("ovp_preis",   "OVP (€)"),
    ("ovp_link",    "Direktlink Anbieter"),
    ("notiz",       "Notiz"),
]
WATCHLIST_FIELDS = [f for f, _ in WATCHLIST_COLUMNS]
WATCHLIST_HEADERS = [h for _, h in WATCHLIST_COLUMNS]

DASHBOARD_COLUMNS: list[tuple[str, str]] = [
    ("Event",                           "Event"),
    ("Kategorie",                       "Kategorie"),
    ("Datum",                           "Datum"),
    ("Venue",                           "Venue"),
    ("Stadt",                           "Stadt"),
    ("Venue_normiert",                  "Venue (normiert)"),
    ("Venue_typ",                       "Venue-Typ"),
    ("Venue_kapazität",                 "Venue-Kapazität"),
    ("Gesamt_Anzahl",                   "Gesamt Angebote"),
    ("Privat_Anzahl",                   "Privat Anz."),
    ("Privat_Min",                      "Privat Min €/K"),
    ("Privat_Avg",                      "Privat Avg €/K"),
    ("Privat_Max",                      "Privat Max €/K"),
    ("Haendler_Anzahl",                 "Händler Anz."),
    ("Haendler_Min",                    "Händler Min €/K"),
    ("Haendler_Avg",                    "Händler Avg €/K"),
    ("Haendler_Max",                    "Händler Max €/K"),
    ("OVP",                             "OVP €/K"),
    ("Marge_Haendler_EUR",              "Händler Marge €"),
    ("Marge_Privat_EUR",                "Privat Marge €"),
    ("Marge_Haendler_Pct",              "Händler Marge %"),
    ("Marge_Privat_Pct",               "Privat Marge %"),
    ("Top_Verkaeufer",                  "Top Verkäufer"),
    ("Top_Verkaeufer_Anzahl",           "Top Verk. Anz."),
    ("Confidence_Modal",                "Confidence"),
    ("Vertrieb_Gewerblich_Anteil_Pct",  "Gewerbl. Anteil %"),
]
DASHBOARD_FIELDS  = [f for f, _ in DASHBOARD_COLUMNS]
DASHBOARD_HEADERS = [h for _, h in DASHBOARD_COLUMNS]

# OVP-Felder: einmal gesetzt, nicht mehr überschreiben
OVP_PROTECTED = {"originalpreis_pro_karte", "ovp_quelle", "ausverkauft"}

# ---------------------------------------------------------------------------
# Farben
# ---------------------------------------------------------------------------

COLOR_HEADER     = "1F4E79"  # dunkelblau (Header)
COLOR_HÄNDLER    = "FCE4D6"  # lachs (Händler-Zeilen)
COLOR_PRIVAT     = "FFFFFF"  # weiß (Privatverkäufer)
COLOR_LOW_CONF   = "FFF2CC"  # gelb (confidence=niedrig)
COLOR_ARCHIV     = "D9D9D9"  # grau (Archiv)
COLOR_OVP_GREEN  = "C6EFCE"  # grün  (OVP vom Anbieter bestätigt)
COLOR_OVP_YELLOW = "FFEB9C"  # gelb  (OVP aus Anzeige extrahiert)
COLOR_OVP_RED    = "FFC7CE"  # rot   (kein OVP gefunden)
COLOR_CONF_GREEN = "C6EFCE"  # grün  (confidence=hoch)
COLOR_CONF_YELLOW= "FFEB9C"  # gelb  (confidence=mittel)
COLOR_CONF_RED   = "FFC7CE"  # rot   (confidence=niedrig)

WHITE = "FFFFFF"


def _header_style(cell):
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_header(ws, headers: list[str]):
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        _header_style(cell)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


_COL_WIDTHS = {
    "Willhaben-Link": 45,
    "Event-Name": 35,
    "Venue": 25,
    "Confidence-Grund": 40,
    "Direktlink Anbieter": 45,
    "Notiz": 30,
}


def _auto_width(ws, headers: list[str]):
    for col_idx, name in enumerate(headers, 1):
        width = _COL_WIDTHS.get(name, max(len(name) + 4, 12))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# ID-Index
# ---------------------------------------------------------------------------

def _build_index(ws, id_field_pos: int = 3) -> dict[str, int]:
    """
    Gibt {willhaben_id: row_number} zurück.
    id_field_pos: 1-basierte Spalte der Anzeigen-ID (Standard = Spalte 3 in MAIN).
    """
    index: dict[str, int] = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        val = row[id_field_pos - 1]
        if val is not None:
            index[str(val)] = row_num
    return index


# ---------------------------------------------------------------------------
# Berechnungen
# ---------------------------------------------------------------------------

def _compute_fields(event: dict) -> dict:
    """Berechnet abgeleitete Felder + Enrichment (vertrieb_klasse, venue_normiert)."""
    from enrichment.vertrieb_erkenner import classify as _classify_vertrieb
    from enrichment.venue_lookup import lookup as _lookup_venue

    result = dict(event)

    gesamt = event.get("angebotspreis_gesamt")
    anzahl = event.get("anzahl_karten")
    pipc = event.get("preis_ist_pro_karte")
    ovp = event.get("originalpreis_pro_karte")

    # Angebotspreis pro Karte
    pro_karte = None
    # angebotspreis_gesamt ist IMMER der Gesamtbetrag für alle Karten.
    # angebotspreis_pro_karte = Gesamtbetrag / Anzahl (unabhängig von pipc).
    if gesamt is not None and anzahl and anzahl > 0:
        pro_karte = round(gesamt / anzahl, 2)
    elif gesamt is not None and pipc is True:
        # Händler mit unbekannter Stückzahl: Gesamtbetrag IST der Preis pro Karte
        pro_karte = gesamt
    # pipc=None oder gesamt=None ohne Anzahl → None
    result["angebotspreis_pro_karte"] = pro_karte

    # Marge
    if pro_karte is not None and ovp is not None and ovp > 0:
        result["marge_eur"] = round(pro_karte - ovp, 2)
        result["marge_pct"] = round((pro_karte - ovp) / ovp * 100, 1)
    else:
        result["marge_eur"] = None
        result["marge_pct"] = None

    # review_nötig
    result["review_nötig"] = "ja" if event.get("confidence") == "niedrig" else "nein"

    # Sprint-1 Enrichment
    result["vertrieb_klasse"] = _classify_vertrieb(event)
    venue_data = _lookup_venue(event.get("venue"))
    result["venue_normiert"]  = venue_data["venue_normiert"]
    result["venue_kapazität"] = venue_data["venue_kapazität"]
    result["venue_typ"]       = venue_data["venue_typ"]

    return result


# ---------------------------------------------------------------------------
# Zeile schreiben
# ---------------------------------------------------------------------------

def _write_row(ws, row_num: int, fields: list[str], data: dict, preserve_ovp: bool = True):
    """Schreibt Daten in eine Zeile. OVP-Felder nur schreiben wenn noch leer."""
    for col_idx, field in enumerate(fields, 1):
        if preserve_ovp and field in OVP_PROTECTED:
            existing = ws.cell(row=row_num, column=col_idx).value
            if existing not in (None, ""):
                continue
        value = data.get(field)
        if isinstance(value, bool):
            value = "ja" if value else "nein"
        if value is None:
            value = ""
        ws.cell(row=row_num, column=col_idx, value=value)


def _apply_cell_colors(ws, row_num: int, data: dict, fields: list[str]):
    """Setzt farbige Zellen für OVP-Quelle und Confidence."""
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=row_num, column=col_idx)

        if field == "originalpreis_pro_karte":
            ovp_quelle = data.get("ovp_quelle", "")
            if ovp_quelle in ("oeticket", "myticket", "konzerthaus"):
                cell.fill = PatternFill("solid", fgColor=COLOR_OVP_GREEN)
            elif ovp_quelle == "Anzeige":
                cell.fill = PatternFill("solid", fgColor=COLOR_OVP_YELLOW)
            elif not data.get("originalpreis_pro_karte"):
                cell.fill = PatternFill("solid", fgColor=COLOR_OVP_RED)

        elif field == "confidence":
            conf = data.get("confidence", "")
            if conf == "hoch":
                cell.fill = PatternFill("solid", fgColor=COLOR_CONF_GREEN)
            elif conf == "mittel":
                cell.fill = PatternFill("solid", fgColor=COLOR_CONF_YELLOW)
            elif conf == "niedrig":
                cell.fill = PatternFill("solid", fgColor=COLOR_CONF_RED)

        # Händler-Zeilen: leichter Lachs-Hintergrund (nur nicht-farbige Zellen)
        elif data.get("verkäufertyp", "").lower() == "händler":
            if cell.fill.fgColor.rgb in ("00000000", WHITE):
                cell.fill = PatternFill("solid", fgColor=COLOR_HÄNDLER)


# ---------------------------------------------------------------------------
# Workbook initialisieren
# ---------------------------------------------------------------------------

def _init_workbook(path: Path) -> Workbook:
    wb = Workbook()

    # 1. Dashboard (index 0 = aktives Sheet beim Öffnen)
    ws_dash = wb.active
    ws_dash.title = SHEET_DASHBOARD
    _write_header(ws_dash, DASHBOARD_HEADERS)
    _auto_width(ws_dash, DASHBOARD_HEADERS)

    # 2. Hauptübersicht
    ws_main = wb.create_sheet(SHEET_HAUPT)
    _write_header(ws_main, MAIN_HEADERS)
    _auto_width(ws_main, MAIN_HEADERS)

    # 3. Review Queue
    ws_review = wb.create_sheet(SHEET_REVIEW)
    _write_header(ws_review, REVIEW_HEADERS)
    _auto_width(ws_review, REVIEW_HEADERS)

    # 4. Watchlist-Config
    ws_watch = wb.create_sheet(SHEET_WATCHLIST)
    _write_header(ws_watch, WATCHLIST_HEADERS)
    _auto_width(ws_watch, WATCHLIST_HEADERS)
    ws_watch.append(["Linkin Park Wien 09.06.2026", 89.90,
                     "https://www.oeticket.com/event/linkin-park-wien-12345", "Beispiel"])

    # 5. Alte Veranstaltungen (letztes Sheet)
    ws_archiv = wb.create_sheet(SHEET_ARCHIV)
    _write_header(ws_archiv, ARCHIV_HEADERS)
    _auto_width(ws_archiv, ARCHIV_HEADERS)

    wb.save(path)
    return wb


def _load_or_create(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    return _init_workbook(path)


# ---------------------------------------------------------------------------
# Haupt-API
# ---------------------------------------------------------------------------

def upsert_events(events: list[dict], excel_path: Path) -> dict:
    """
    Schreibt/aktualisiert Events in die Excel-Datei.
    Rückgabe: {inserted, updated, review_added}
    """
    wb = _load_or_create(excel_path)
    ws_main   = wb[SHEET_HAUPT]
    ws_review = wb[SHEET_REVIEW]

    # Anzeigen-ID ist in Spalte 3 (MAIN_FIELDS[2] = "willhaben_id")
    main_index = _build_index(ws_main, id_field_pos=3)
    # Review-Queue: Anzeigen-ID ist Spalte 1 (REVIEW_FIELDS[0] = "willhaben_id")
    review_index = _build_index(ws_review, id_field_pos=1)

    stats = {"inserted": 0, "updated": 0, "review_added": 0}

    for event in events:
        wid = str(event.get("willhaben_id", "")).strip()
        if not wid:
            continue

        # Abgeleitete Felder berechnen
        event = _compute_fields(event)
        event.setdefault("scan_datum", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
        event.setdefault("watchlist", "nein")

        # --- Hauptübersicht ---
        if wid in main_index:
            row_num = main_index[wid]
            _write_row(ws_main, row_num, MAIN_FIELDS, event, preserve_ovp=True)
            _apply_cell_colors(ws_main, row_num, event, MAIN_FIELDS)
            stats["updated"] += 1
        else:
            next_row = ws_main.max_row + 1
            _write_row(ws_main, next_row, MAIN_FIELDS, event, preserve_ovp=False)
            _apply_cell_colors(ws_main, next_row, event, MAIN_FIELDS)
            main_index[wid] = next_row
            stats["inserted"] += 1

        # --- Review Queue: nur confidence=niedrig ---
        if event.get("confidence") == "niedrig" and wid not in review_index:
            next_row = ws_review.max_row + 1
            _write_row(ws_review, next_row, REVIEW_FIELDS, event, preserve_ovp=False)
            review_index[wid] = next_row
            stats["review_added"] += 1

    wb.save(excel_path)
    return stats


def _read_row_as_dict(ws, row_num: int) -> dict:
    """Liest eine Hauptübersicht-Zeile in ein MAIN_FIELDS-keyiertes dict."""
    return {
        field: ws.cell(row=row_num, column=col).value
        for col, field in enumerate(MAIN_FIELDS, start=1)
    }


def update_hauptuebersicht_mit_historie(
    events: list[dict],
    excel_path: Path,
    scan_datum: "datetime.date | None" = None,
) -> dict:
    """Append-Logik: merged neue Scrape-Daten in Hauptübersicht unter Erhalt der Preis-Historie.

    Statt blindem Überschreiben (upsert_events) wird die bestehende Zeile zuerst gelesen
    und mit HistorieManager gemerged. Neue Zeilen erhalten initialisierte History-Felder.
    Rückgabe: {inserted, updated, review_added}.
    """
    from parser.v2.historie_manager import merge_scrape_mit_historie, markiere_inaktive

    if scan_datum is None:
        scan_datum = datetime.date.today()

    wb = _load_or_create(excel_path)
    ws_haupt  = wb[SHEET_HAUPT]
    ws_review = wb[SHEET_REVIEW]

    main_index   = _build_index(ws_haupt, id_field_pos=3)
    review_index = _build_index(ws_review, id_field_pos=1)

    current_ids = {str(e["willhaben_id"]) for e in events if e.get("willhaben_id")}
    stats = {"inserted": 0, "updated": 0, "review_added": 0}

    for event in events:
        event = _compute_fields(event)
        event.setdefault("scan_datum", scan_datum.strftime("%Y-%m-%d"))
        event.setdefault("watchlist", "nein")
        wid = str(event.get("willhaben_id") or "").strip()
        if not wid:
            continue

        if wid in main_index:
            row_num  = main_index[wid]
            existing = _read_row_as_dict(ws_haupt, row_num)
            # None/"" aus openpyxl normalisieren: leere Strings als None behandeln
            existing = {k: (v if v != "" else None) for k, v in existing.items()}
            merged   = merge_scrape_mit_historie(existing, event, scan_datum)
            _write_row(ws_haupt, row_num, MAIN_FIELDS, merged, preserve_ovp=True)
            _apply_cell_colors(ws_haupt, row_num, merged, MAIN_FIELDS)
            stats["updated"] += 1
        else:
            event["erstmals_gesehen"]         = scan_datum.isoformat()
            event["zuletzt_gesehen"]          = scan_datum.isoformat()
            event["status"]                   = "aktiv"
            event["scan_anzahl"]              = 1
            event["preis_aktuell"]            = event.get("angebotspreis_pro_karte")
            event["preis_vor_7_tagen"]        = None
            event["preis_aenderungen_count"]  = 0
            event["letzte_preisaenderung_am"] = scan_datum.isoformat()
            next_row = ws_haupt.max_row + 1
            _write_row(ws_haupt, next_row, MAIN_FIELDS, event, preserve_ovp=False)
            _apply_cell_colors(ws_haupt, next_row, event, MAIN_FIELDS)
            main_index[wid] = next_row
            stats["inserted"] += 1

        if event.get("confidence") == "niedrig" and wid not in review_index:
            next_row = ws_review.max_row + 1
            _write_row(ws_review, next_row, REVIEW_FIELDS, event, preserve_ovp=False)
            review_index[wid] = next_row
            stats["review_added"] += 1

    # Inaktiv-Markierung: alle Zeilen prüfen
    id_col = MAIN_FIELDS.index("willhaben_id") + 1
    status_col = MAIN_FIELDS.index("status") + 1
    all_rows = [
        _read_row_as_dict(ws_haupt, r)
        for r in range(2, ws_haupt.max_row + 1)
        if ws_haupt.cell(row=r, column=id_col).value not in (None, "")
    ]
    updated_rows = markiere_inaktive(all_rows, current_ids, scan_datum)
    for i, row_dict in enumerate(updated_rows):
        data_row = i + 2
        ws_haupt.cell(row=data_row, column=status_col).value = row_dict["status"]

    wb.save(excel_path)
    return stats


def write_dashboard(agg_rows: list[dict], excel_path: Path) -> None:
    """
    Schreibt (oder überschreibt) das Dashboard-Sheet mit aggregierten Daten.
    agg_rows ist eine Liste von Dicts (aus aggregate().to_dict('records')).
    """
    wb = _load_or_create(excel_path)

    if SHEET_DASHBOARD in wb.sheetnames:
        del wb[SHEET_DASHBOARD]
    ws = wb.create_sheet(SHEET_DASHBOARD, 0)

    _write_header(ws, DASHBOARD_HEADERS)
    _auto_width(ws, DASHBOARD_HEADERS)

    for row in agg_rows:
        vals = []
        for field in DASHBOARD_FIELDS:
            val = row.get(field)
            if isinstance(val, float) and math.isnan(val):
                val = ""
            elif val is None:
                val = ""
            elif isinstance(val, float):
                val = round(val, 2)
            vals.append(val)
        ws.append(vals)

    wb.save(excel_path)


def finalisiere_lauf(events: list[dict], excel_path: Path) -> dict:
    """
    Vollständiger Pipeline-Schritt: upsert → archivierung → Dashboard-Refresh.
    Rückgabe: {inserted, updated, review_added, archived}.
    """
    from export.archivierung import archive_expired as _archive
    from app.backend.dashboard_aggregator import load_excel, aggregate

    excel_stats = update_hauptuebersicht_mit_historie(events, excel_path)
    archived = _archive(excel_path)

    df = load_excel(excel_path)
    agg = aggregate(df)
    agg_rows = agg.to_dict("records") if not agg.empty else []
    write_dashboard(agg_rows, excel_path)

    return {**excel_stats, "archived": archived}


def update_ovp(
    excel_path: Path,
    willhaben_id: str,
    originalpreis_pro_karte: float | None,
    ovp_quelle: str,
    ausverkauft: str,
) -> bool:
    """
    Aktualisiert OVP-Felder für eine Willhaben-ID.
    Gibt True zurück wenn die ID gefunden wurde.
    """
    wb = _load_or_create(excel_path)
    ws = wb[SHEET_HAUPT]
    index = _build_index(ws, id_field_pos=3)

    if willhaben_id not in index:
        return False

    row_num = index[willhaben_id]
    ovp_col     = MAIN_FIELDS.index("originalpreis_pro_karte") + 1
    quelle_col  = MAIN_FIELDS.index("ovp_quelle") + 1
    ausvk_col   = MAIN_FIELDS.index("ausverkauft") + 1

    if originalpreis_pro_karte is not None:
        ws.cell(row=row_num, column=ovp_col).value = originalpreis_pro_karte
    ws.cell(row=row_num, column=quelle_col).value = ovp_quelle
    ws.cell(row=row_num, column=ausvk_col).value = ausverkauft

    # Farbe der OVP-Zelle aktualisieren
    ovp_cell = ws.cell(row=row_num, column=ovp_col)
    if ovp_quelle in ("oeticket", "myticket", "konzerthaus"):
        ovp_cell.fill = PatternFill("solid", fgColor=COLOR_OVP_GREEN)
    elif ovp_quelle == "Anzeige":
        ovp_cell.fill = PatternFill("solid", fgColor=COLOR_OVP_YELLOW)

    wb.save(excel_path)
    return True


if __name__ == "__main__":
    import tempfile
    path = Path(tempfile.mkdtemp()) / "test_willhaben.xlsx"
    sample_events = [
        {
            "willhaben_id": "2123545129",
            "willhaben_link": "https://www.willhaben.at/iad/kaufen-und-verkaufen/d/linkin-park-wien-2123545129/",
            "scan_datum": "2026-04-14 10:00",
            "verkäufer_id": "987654",
            "verkäufername": "TicketsWien",
            "verkäufertyp": "Händler",
            "mitglied_seit": "03/2018",
            "event_name": "Linkin Park",
            "event_datum": "2026-06-09",
            "venue": "Ernst Happel Stadion",
            "stadt": "Wien",
            "kategorie": "Front-of-Stage",
            "anzahl_karten": 1,
            "angebotspreis_gesamt": 250.0,
            "preis_ist_pro_karte": True,
            "originalpreis_pro_karte": 89.90,
            "ovp_quelle": "Anzeige",
            "ausverkauft": "unbekannt",
            "watchlist": "ja",
            "confidence": "hoch",
            "confidence_grund": None,
        },
        {
            "willhaben_id": "9999999",
            "willhaben_link": "https://www.willhaben.at/iad/kaufen-und-verkaufen/d/bad-bunny-9999999/",
            "verkäufer_id": "111",
            "verkäufername": "Anna",
            "verkäufertyp": "Privat",
            "mitglied_seit": "07/2022",
            "event_name": "Bad Bunny",
            "event_datum": "2026-07-20",
            "venue": None,
            "stadt": "Stockholm",
            "kategorie": "Unbekannt",
            "anzahl_karten": None,
            "angebotspreis_gesamt": 240.0,
            "preis_ist_pro_karte": None,
            "originalpreis_pro_karte": None,
            "ovp_quelle": "",
            "ausverkauft": "unbekannt",
            "confidence": "niedrig",
            "confidence_grund": "Preis unklar ob pro Karte oder gesamt",
        },
    ]
    stats = upsert_events(sample_events, path)
    print(f"Gespeichert: {path}")
    print(f"Stats: {stats}")
    archived = archive_expired(path, cutoff_date=datetime.date(2026, 5, 1))
    print(f"Archiviert: {archived}")
