"""Archive expired events from Hauptübersicht → Archiv sheet."""
from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from export.excel_writer import MAIN_FIELDS, COLOR_ARCHIV, SHEET_HAUPT, SHEET_ARCHIV


def archive_expired(excel_path: Path, cutoff_date: datetime.date | None = None) -> int:
    """
    Copies rows with event_datum < cutoff_date from Hauptübersicht to Archiv,
    sets archiviert_am = today (ISO), grays them out in the main sheet.
    Returns number of archived rows.
    """
    if not excel_path.exists():
        return 0
    if cutoff_date is None:
        cutoff_date = datetime.date.today() - datetime.timedelta(days=30)

    wb = load_workbook(excel_path)
    ws_main   = wb[SHEET_HAUPT]
    ws_archiv = wb[SHEET_ARCHIV]

    datum_col = MAIN_FIELDS.index("event_datum") + 1
    archiviert_am_col = MAIN_FIELDS.index("archiviert_am") + 1
    today_str = datetime.date.today().isoformat()
    n_cols = len(MAIN_FIELDS)

    rows_to_archive: list[int] = []
    for row_num in range(2, ws_main.max_row + 1):
        val = ws_main.cell(row=row_num, column=datum_col).value
        if not val:
            continue
        try:
            if isinstance(val, str):
                event_date = datetime.date.fromisoformat(val[:10])
            elif isinstance(val, datetime.datetime):
                event_date = val.date()
            elif isinstance(val, datetime.date):
                event_date = val
            else:
                continue
            if event_date < cutoff_date:
                rows_to_archive.append(row_num)
        except ValueError:
            continue

    archived = 0
    for row_num in sorted(rows_to_archive, reverse=True):
        row_vals = [ws_main.cell(row=row_num, column=c).value for c in range(1, n_cols + 1)]
        row_vals[archiviert_am_col - 1] = today_str
        ws_archiv.append(row_vals)

        for col_idx in range(1, n_cols + 1):
            ws_main.cell(row=row_num, column=col_idx).fill = PatternFill(
                "solid", fgColor=COLOR_ARCHIV
            )
        archived += 1

    if archived:
        wb.save(excel_path)

    return archived


def archive_aeltere_als(excel_path: Path, tage: int = 30) -> int:
    """
    Archives rows where event_datum < heute - tage.
    Convenience wrapper around archive_expired.
    """
    cutoff = datetime.date.today() - datetime.timedelta(days=tage)
    return archive_expired(excel_path, cutoff_date=cutoff)
