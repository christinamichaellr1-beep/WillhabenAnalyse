"""CLI: verify event data in Excel against external APIs.

Usage:
    python -m verification.verify_excel [--excel PATH] [--limit N] [--dry-run]
"""
from __future__ import annotations
import argparse
import datetime
import logging
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

from export.excel_writer import MAIN_FIELDS, SHEET_HAUPT
from verification.cache import VerificationCache
from verification.excel_integration import write_verif_result, rebuild_nicht_verifiziert_sheet
from verification.orchestrator import Orchestrator, VerifStatus

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

DEFAULT_EXCEL = Path("data/willhaben_markt.xlsx")
DEFAULT_DB    = Path("data/verification_cache.db")


def _read_hauptuebersicht(excel_path: Path) -> list[dict]:
    """Read all data rows from Hauptübersicht as list of field dicts."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if SHEET_HAUPT not in wb.sheetnames:
        return []
    ws = wb[SHEET_HAUPT]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        d = {field: row[i] for i, field in enumerate(MAIN_FIELDS) if i < len(row)}
        rows.append(d)
    wb.close()
    return rows


def run(
    excel_path: Path,
    db_path: Path,
    ttl_days: int = 7,
    limit: int | None = None,
    dry_run: bool = False,
    force: bool = False,
    status_filter: str | None = None,
) -> dict:
    """
    Main verification loop.
    Returns stats dict: {total, cached, verified, likely, unverified, failed, skipped, errors}.
    """
    if not excel_path.exists():
        logger.error("Excel nicht gefunden: %s", excel_path)
        return {}

    cache = VerificationCache(db_path=db_path, ttl_days=ttl_days)
    orchestrator = Orchestrator()
    rows = _read_hauptuebersicht(excel_path)

    # Filter by status if requested
    if status_filter:
        rows = [r for r in rows if str(r.get("verif_status") or "").strip() == status_filter]

    if limit:
        rows = rows[:limit]

    stats: dict[str, int] = {
        "total": len(rows), "cached": 0,
        "verifiziert": 0, "wahrscheinlich": 0,
        "nicht_verifiziert": 0, "fehler": 0, "übersprungen": 0, "errors": 0,
    }

    for i, row in enumerate(rows, 1):
        wid = str(row.get("willhaben_id") or "").strip()
        event_name = str(row.get("event_name") or "").strip() or None
        stadt = str(row.get("stadt") or "").strip() or None

        # Parse event_datum
        datum_raw = row.get("event_datum")
        event_datum = None
        if datum_raw:
            try:
                if isinstance(datum_raw, datetime.date):
                    event_datum = datum_raw
                else:
                    event_datum = datetime.date.fromisoformat(str(datum_raw)[:10])
            except ValueError:
                pass

        logger.info("[%d/%d] Verifying: %s (ID=%s)", i, len(rows), event_name or "?", wid)

        # Cache check
        if not force and event_name:
            cached = cache.get(event_name, event_datum)
            if cached:
                stats["cached"] += 1
                stats[cached.status.value] = stats.get(cached.status.value, 0) + 1
                if not dry_run and wid:
                    write_verif_result(excel_path, wid, cached)
                logger.info("  -> Cache: %s", cached.status.value)
                continue

        # Verify
        try:
            result = orchestrator.verify(event_name, event_datum, stadt)
        except Exception as exc:
            logger.error("  Fehler bei Verifikation: %s", exc)
            stats["errors"] += 1
            continue

        stats[result.status.value] = stats.get(result.status.value, 0) + 1

        if event_name:
            cache.put(event_name, event_datum, result)

        if not dry_run and wid:
            try:
                write_verif_result(excel_path, wid, result)
            except Exception as exc:
                logger.error("  Excel-Schreibfehler: %s", exc)
                stats["errors"] += 1

        logger.info(
            "  -> %s (score=%s)", result.status.value,
            round(result.best_match.total_score, 3) if result.best_match else "n/a",
        )

    if not dry_run:
        nv_count = rebuild_nicht_verifiziert_sheet(excel_path)
        logger.info("Nicht-Verifiziert Sheet: %d Zeilen", nv_count)

    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Verify Willhaben events against external APIs")
    parser.add_argument("--excel",    default=str(DEFAULT_EXCEL), help="Pfad zur Excel-Datei")
    parser.add_argument("--db",       default=str(DEFAULT_DB),    help="SQLite Cache Pfad")
    parser.add_argument("--ttl-days", type=int, default=7,         help="Cache TTL in Tagen")
    parser.add_argument("--limit",    type=int, default=None,      help="Max Zeilen")
    parser.add_argument("--dry-run",  action="store_true",          help="Keine Änderungen schreiben")
    parser.add_argument("--force",    action="store_true",          help="Cache ignorieren")
    parser.add_argument("--status",   default=None,                 help="Nur Zeilen mit diesem Status")
    args = parser.parse_args()

    stats = run(
        excel_path=Path(args.excel),
        db_path=Path(args.db),
        ttl_days=args.ttl_days,
        limit=args.limit,
        dry_run=args.dry_run,
        force=args.force,
        status_filter=args.status,
    )
    print("\n=== Verifikations-Ergebnis ===")
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
