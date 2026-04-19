# verification/excel_integration.py
"""Write VerificationResult back into the Excel file."""
from __future__ import annotations
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from export.excel_writer import (
    MAIN_FIELDS, MAIN_HEADERS, SHEET_HAUPT,
    SHEET_NICHT_VERIFIZIERT, _write_header, _auto_width,
)
from verification.orchestrator import VerifStatus, VerificationResult

logger = logging.getLogger(__name__)

# Color coding for verif_status cells
COLOR_VERIFIED   = "C6EFCE"   # grün
COLOR_LIKELY     = "FFEB9C"   # gelb
COLOR_UNVERIFIED = "FFC7CE"   # rot
COLOR_FAILED     = "D9D9D9"   # grau
COLOR_SKIPPED    = "FFFFFF"   # weiß

_STATUS_COLORS = {
    VerifStatus.VERIFIED:   COLOR_VERIFIED,
    VerifStatus.LIKELY:     COLOR_LIKELY,
    VerifStatus.UNVERIFIED: COLOR_UNVERIFIED,
    VerifStatus.FAILED:     COLOR_FAILED,
    VerifStatus.SKIPPED:    COLOR_SKIPPED,
}

# Column indices (1-based) for verification fields
_VERIF_STATUS_COL = MAIN_FIELDS.index("verif_status") + 1
_VERIF_QUELLEN_COL = MAIN_FIELDS.index("verif_quellen") + 1
_VERIF_DATUM_COL = MAIN_FIELDS.index("verif_datum") + 1
_VERIF_NAME_COL = MAIN_FIELDS.index("verif_name") + 1
_VERIF_SCORE_COL = MAIN_FIELDS.index("verif_score") + 1
_ID_COL = MAIN_FIELDS.index("willhaben_id") + 1


def _result_to_verif_fields(result: VerificationResult) -> dict:
    """Convert VerificationResult to the 5 verif_ field values."""
    score = None
    name = None
    if result.best_match:
        score = round(result.best_match.total_score, 3)
        name = result.best_match.candidate.event_name
    return {
        "verif_status": result.status.value,
        "verif_quellen": "; ".join(result.sources_confirmed),
        "verif_datum": result.verif_datum,
        "verif_name": name or "",
        "verif_score": score,
    }


def write_verif_result(
    excel_path: Path,
    willhaben_id: str,
    result: VerificationResult,
) -> bool:
    """
    Write VerificationResult into the Hauptübersicht row for the given willhaben_id.
    Colors the verif_status cell. Returns True if row found and updated, False otherwise.
    """
    wb = openpyxl.load_workbook(excel_path)
    if SHEET_HAUPT not in wb.sheetnames:
        logger.warning("Sheet %s nicht gefunden", SHEET_HAUPT)
        return False

    ws = wb[SHEET_HAUPT]
    fields = _result_to_verif_fields(result)

    for row_num in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row_num, column=_ID_COL).value
        if str(cell_id).strip() == str(willhaben_id).strip():
            ws.cell(row=row_num, column=_VERIF_STATUS_COL).value = fields["verif_status"]
            ws.cell(row=row_num, column=_VERIF_QUELLEN_COL).value = fields["verif_quellen"]
            ws.cell(row=row_num, column=_VERIF_DATUM_COL).value = fields["verif_datum"]
            ws.cell(row=row_num, column=_VERIF_NAME_COL).value = fields["verif_name"]
            ws.cell(row=row_num, column=_VERIF_SCORE_COL).value = fields["verif_score"]
            # Color status cell
            color = _STATUS_COLORS.get(result.status, COLOR_SKIPPED)
            ws.cell(row=row_num, column=_VERIF_STATUS_COL).fill = PatternFill("solid", fgColor=color)
            wb.save(excel_path)
            return True

    logger.warning("willhaben_id %s nicht in Hauptübersicht gefunden", willhaben_id)
    return False


def rebuild_nicht_verifiziert_sheet(excel_path: Path) -> int:
    """
    (Re-)creates the 'Nicht-Verifiziert' sheet with all rows from Hauptübersicht
    where verif_status is empty, None, or 'nicht_verifiziert'.
    Returns count of rows written.
    """
    wb = openpyxl.load_workbook(excel_path)
    if SHEET_HAUPT not in wb.sheetnames:
        return 0

    ws_haupt = wb[SHEET_HAUPT]

    # Remove and recreate the sheet
    if SHEET_NICHT_VERIFIZIERT in wb.sheetnames:
        del wb[SHEET_NICHT_VERIFIZIERT]
    ws_nv = wb.create_sheet(SHEET_NICHT_VERIFIZIERT)
    _write_header(ws_nv, MAIN_HEADERS)
    _auto_width(ws_nv, MAIN_HEADERS)

    verif_status_col = _VERIF_STATUS_COL
    n_cols = len(MAIN_FIELDS)
    count = 0

    for row_num in range(2, ws_haupt.max_row + 1):
        wid = ws_haupt.cell(row=row_num, column=_ID_COL).value
        if wid is None:
            continue
        status_val = ws_haupt.cell(row=row_num, column=verif_status_col).value
        is_unverified = (
            not status_val
            or str(status_val).strip() == ""
            or status_val == VerifStatus.UNVERIFIED.value
        )
        if is_unverified:
            row_vals = [ws_haupt.cell(row=row_num, column=c).value for c in range(1, n_cols + 1)]
            ws_nv.append(row_vals)
            count += 1

    wb.save(excel_path)
    return count
