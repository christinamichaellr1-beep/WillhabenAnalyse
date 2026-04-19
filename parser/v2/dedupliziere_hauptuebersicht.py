"""
dedupliziere_hauptuebersicht.py

Standalone-Modul zum Deduplizieren des Hauptübersicht-Sheets in einer Excel-Datei.
Identifiziert Zeilen mit doppelter willhaben_id, behält die LETZTE (höchste Zeilennummer),
entfernt frühere Duplikate. Gibt Anzahl der entfernten Zeilen zurück.
"""
from pathlib import Path

import openpyxl

from export.excel_writer import MAIN_FIELDS, SHEET_HAUPT


def dedupliziere_hauptuebersicht(excel_path: Path) -> int:
    """
    Removes duplicate rows from Hauptübersicht sheet (by willhaben_id).
    Keeps the LAST occurrence of each willhaben_id (highest row number).
    Rewrites the sheet in-place.
    Returns number of removed rows.
    """
    if not excel_path.exists():
        return 0

    wb = openpyxl.load_workbook(excel_path)
    if SHEET_HAUPT not in wb.sheetnames:
        return 0

    ws = wb[SHEET_HAUPT]
    id_col = MAIN_FIELDS.index("willhaben_id") + 1  # 1-based

    # Collect all rows with their willhaben_id
    rows: list[tuple[int, str, list]] = []  # (row_num, wid, values)
    for row_num in range(2, ws.max_row + 1):
        wid_val = ws.cell(row=row_num, column=id_col).value
        if wid_val is None:
            continue
        wid = str(wid_val).strip()
        if not wid:
            continue
        values = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
        rows.append((row_num, wid, values))

    # Find duplicates — keep last occurrence per id
    seen: dict[str, int] = {}  # wid → index in rows
    for i, (_, wid, _) in enumerate(rows):
        seen[wid] = i

    keep_indices = set(seen.values())
    removed = len(rows) - len(keep_indices)

    if removed == 0:
        return 0

    # Rewrite sheet: clear data rows, then write kept rows in original order
    for row_num in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col).value = None

    for new_row_idx, (_, _, values) in enumerate(
        (rows[i] for i in sorted(keep_indices)), start=2
    ):
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=new_row_idx, column=col_idx).value = val

    wb.save(excel_path)
    return removed
