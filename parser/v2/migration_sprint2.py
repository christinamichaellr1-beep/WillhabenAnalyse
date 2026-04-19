"""
migration_sprint2.py

Fügt die 8 Sprint-2-Historien-Spalten (35–42) mit smarten Defaults
in eine bestehende Hauptübersicht-Workbook ein.

Idempotent: Zeilen, die Spalte 35 bereits befüllt haben, werden übersprungen.

CLI: python -m parser.v2.migration_sprint2 <pfad-zur-excel>
"""
from pathlib import Path

from openpyxl import load_workbook

from export.excel_writer import MAIN_HEADERS, SHEET_HAUPT

# Sprint-2-Spalten starten bei Index 34 (0-basiert) → Excel-Spalte 35 (1-basiert)
_SPRINT2_HEADERS = MAIN_HEADERS[34:]   # 8 Einträge: "Erstmals gesehen" … "Letzte Preisänderung am"
_SPRINT2_START_COL = 35                # 1-basiert in Excel

# Positionen der Quell-Spalten (1-basiert) für smarte Defaults
_COL_SCAN_DATUM           = 1   # "Scan-Datum"
_COL_WILLHABEN_ID         = 3   # "Anzeigen-ID" (PK-Check)
_COL_ANGEBOTSPREIS_PRO_K  = 16  # "Angebotspreis pro Karte"


def migrate_sprint2(excel_path: Path) -> int:
    """Migriert Hauptübersicht: fügt 8 Sprint-2-Spalten ein und setzt Defaults.

    Gibt die Anzahl der migrierten Datenzeilen zurück.
    """
    wb = load_workbook(excel_path)

    if SHEET_HAUPT not in wb.sheetnames:
        return 0

    ws = wb[SHEET_HAUPT]

    # Schritt 1: Header für Spalten 35–42 schreiben
    for i, header in enumerate(_SPRINT2_HEADERS, start=_SPRINT2_START_COL):
        ws.cell(row=1, column=i).value = header

    # Schritt 2: Datenzeilen mit Defaults füllen (idempotent: Spalte 35 gesetzt → skip)
    row_count = 0
    for row_num in range(2, ws.max_row + 1):
        if ws.cell(row=row_num, column=_COL_WILLHABEN_ID).value is None:
            continue  # Leerzeile

        if ws.cell(row=row_num, column=_SPRINT2_START_COL).value not in (None, ""):
            row_count += 1
            continue  # Bereits migriert

        scan_datum_raw  = ws.cell(row=row_num, column=_COL_SCAN_DATUM).value
        preis_pro_karte = ws.cell(row=row_num, column=_COL_ANGEBOTSPREIS_PRO_K).value

        scan_datum_str = str(scan_datum_raw)[:10] if scan_datum_raw else ""

        defaults = [
            scan_datum_str,   # 35: erstmals_gesehen
            scan_datum_str,   # 36: zuletzt_gesehen
            "aktiv",          # 37: status
            1,                # 38: scan_anzahl
            preis_pro_karte,  # 39: preis_aktuell (aus vorhandenem Preis)
            None,             # 40: preis_vor_7_tagen
            0,                # 41: preis_aenderungen_count
            "",               # 42: letzte_preisaenderung_am
        ]
        for i, val in enumerate(defaults, start=_SPRINT2_START_COL):
            ws.cell(row=row_num, column=i).value = val

        row_count += 1

    wb.save(excel_path)
    return row_count


if __name__ == "__main__":
    import sys

    path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/willhaben_markt.xlsx")
    count = migrate_sprint2(path)
    print(f"Migriert: {count} Zeilen in {path}")
