"""
cleanup_dashboard_muell.py

Identifiziert Müll-Zeilen in der Hauptübersicht (fehlender Event-Name,
fehlendes Datum, kein Preis) und verschiebt sie in ein neues Sheet
"Gefilterter Müll". Anschließend wird das Dashboard neu generiert.

Idempotent: Bereits im Müll-Sheet vorhandene Zeilen werden nicht dupliziert.
Bestehende Müll-Zeilen aus vorherigen Läufen werden beibehalten.

CLI: python -m scripts.cleanup_dashboard_muell [--excel <pfad>] [--dry-run]
"""
from __future__ import annotations

import argparse
import logging
from pathlib import Path

import openpyxl
import pandas as pd

from app.backend.dashboard_aggregator import filtere_dashboard_input, load_excel
from export.excel_writer import MAIN_FIELDS, SHEET_HAUPT, write_dashboard

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

SHEET_MUELL = "Gefilterter Müll"

# Willhaben-ID steht in MAIN_FIELDS[2] (0-basiert) → Excel-Spalte 3
_ID_COL_1_BASED = MAIN_FIELDS.index("willhaben_id") + 1


def _collect_muell_ids_in_sheet(ws_muell: openpyxl.worksheet.worksheet.Worksheet) -> set[str]:
    """Liest alle Willhaben-IDs aus dem Müll-Sheet (Spalte _ID_COL_1_BASED)."""
    ids: set[str] = set()
    for row in range(2, ws_muell.max_row + 1):
        val = ws_muell.cell(row=row, column=_ID_COL_1_BASED).value
        if val not in (None, ""):
            ids.add(str(val))
    return ids


def _copy_row(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        dst_ws.cell(row=dst_row, column=col).value = src_ws.cell(row=src_row, column=col).value


def cleanup_muell(excel_path: Path, dry_run: bool = False) -> dict:
    """Verschiebt Müll-Zeilen aus Hauptübersicht in 'Gefilterter Müll' Sheet.

    Gibt {'moved': int, 'kept': int} zurück.
    """
    wb = openpyxl.load_workbook(excel_path)

    if SHEET_HAUPT not in wb.sheetnames:
        logger.warning("Sheet '%s' nicht gefunden — nichts zu tun.", SHEET_HAUPT)
        return {"moved": 0, "kept": 0}

    ws_haupt = wb[SHEET_HAUPT]
    max_col = ws_haupt.max_column

    # Müll-Sheet anlegen falls nicht vorhanden
    if SHEET_MUELL not in wb.sheetnames:
        wb.create_sheet(SHEET_MUELL)
        ws_muell = wb[SHEET_MUELL]
        # Header kopieren
        for col in range(1, max_col + 1):
            ws_muell.cell(row=1, column=col).value = ws_haupt.cell(row=1, column=col).value
    else:
        ws_muell = wb[SHEET_MUELL]

    already_moved: set[str] = _collect_muell_ids_in_sheet(ws_muell)
    muell_next_row = ws_muell.max_row + 1

    # Alle Datenzeilen aus Hauptübersicht lesen und klassifizieren
    haupt_rows: list[tuple[int, dict]] = []
    for row_num in range(2, ws_haupt.max_row + 1):
        wid = ws_haupt.cell(row=row_num, column=_ID_COL_1_BASED).value
        if wid is None:
            continue
        row_dict = {
            field: ws_haupt.cell(row=row_num, column=i + 1).value
            for i, field in enumerate(MAIN_FIELDS)
            if i + 1 <= max_col
        }
        haupt_rows.append((row_num, row_dict))

    if not haupt_rows:
        logger.info("Keine Datenzeilen in Hauptübersicht gefunden.")
        return {"moved": 0, "kept": 0}

    # DataFrame für filtere_dashboard_input aufbauen
    df_all = pd.DataFrame([r for _, r in haupt_rows])
    df_sauber = filtere_dashboard_input(df_all)
    sauber_ids: set[str] = set(df_sauber["willhaben_id"].astype(str))

    moved = 0
    keep_rows: list[int] = []

    for row_num, row_dict in haupt_rows:
        wid = str(row_dict.get("willhaben_id", ""))
        if wid in sauber_ids:
            keep_rows.append(row_num)
        else:
            # Müll — in Müll-Sheet verschieben (nur wenn nicht schon da)
            if wid not in already_moved:
                if not dry_run:
                    _copy_row(ws_haupt, ws_muell, row_num, muell_next_row, max_col)
                    muell_next_row += 1
                    already_moved.add(wid)
                moved += 1
                logger.info("Müll: %s — %s", wid, row_dict.get("event_name"))

    kept = len(keep_rows)
    logger.info("Ergebnis: %d sauber behalten, %d als Müll markiert.", kept, moved)

    if dry_run:
        logger.info("--dry-run: keine Änderungen gespeichert.")
        return {"moved": moved, "kept": kept}

    # Hauptübersicht neu aufbauen: nur saubere Zeilen (in-place via Clear + Rewrite)
    # Bestehende Datenzeilen löschen (Zeile 2 bis max_row)
    for row_num in range(2, ws_haupt.max_row + 1):
        for col in range(1, max_col + 1):
            ws_haupt.cell(row=row_num, column=col).value = None

    # Saubere Zeilen neu schreiben
    for new_row, row_dict in enumerate(
        (r for _, r in haupt_rows if str(r.get("willhaben_id", "")) in sauber_ids),
        start=2,
    ):
        for i, field in enumerate(MAIN_FIELDS):
            if i + 1 <= max_col:
                ws_haupt.cell(row=new_row, column=i + 1).value = row_dict.get(field)

    wb.save(excel_path)
    logger.info("Gespeichert: %s", excel_path)
    return {"moved": moved, "kept": kept}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Dashboard-Müll bereinigen")
    parser.add_argument("--excel", default="data/willhaben_markt.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Keine Änderungen speichern")
    parser.add_argument("--regenerate-dashboard", action="store_true",
                        help="Dashboard nach Bereinigung neu generieren")
    args = parser.parse_args()

    target = Path(args.excel)
    if not target.exists():
        print(f"FEHLER: Datei nicht gefunden: {target}")
        raise SystemExit(1)

    result = cleanup_muell(target, dry_run=args.dry_run)
    print(f"Müll verschoben: {result['moved']}, sauber behalten: {result['kept']}")

    if args.regenerate_dashboard and not args.dry_run:
        from app.backend.dashboard_aggregator import aggregate
        df = load_excel(target)
        df_sauber = filtere_dashboard_input(df)
        agg = aggregate(df_sauber)
        write_dashboard(agg.to_dict("records"), target)
        print(f"Dashboard neu generiert: {len(agg)} Gruppen.")
